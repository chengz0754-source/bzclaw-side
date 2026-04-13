#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import math
import re
import shutil
import sys
import time
import traceback
from collections import Counter
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

SOURCE_TOOL = "sellersprite_market_research"
ROUTE_TYPE = "market"
STAGE_CODE = "M02"

CN_NICHE_EN = "\u7ec6\u5206\u5e02\u573a"
CN_NICHE_ZH = "\u7ec6\u5206\u5e02\u573a(\u7ffb\u8bd1)"
CN_MARKET_PATH = "\u5e02\u573a\u8def\u5f84"
CN_SAMPLE_QUANTITY = "\u6837\u672c\u6570\u91cf"
CN_MONTHLY_TOTAL_SALES = "\u6708\u603b\u9500\u91cf"
CN_MONTHLY_AVG_SALES = "\u6708\u5747\u9500\u91cf"
CN_MONTHLY_AVG_REVENUE = "\u6708\u5747\u9500\u552e\u989d($)"
CN_AVG_PRICE = "\u5e73\u5747\u4ef7\u683c($)"
CN_AVG_REVIEW_COUNT = "\u5e73\u5747\u8bc4\u5206\u6570"
CN_AVG_STAR = "\u5e73\u5747\u661f\u7ea7"
CN_AVG_BSR = "\u5e73\u5747BSR"
CN_AVG_SELLER_COUNT = "\u5e73\u5747\u5356\u5bb6\u6570"
CN_SELLER_TYPE = "\u5356\u5bb6\u7c7b\u578b"
CN_PRODUCT_CONCENTRATION = "\u5546\u54c1\u96c6\u4e2d\u5ea6"
CN_BRAND_CONCENTRATION = "\u54c1\u724c\u96c6\u4e2d\u5ea6"
CN_SELLER_CONCENTRATION = "\u5356\u5bb6\u96c6\u4e2d\u5ea6"
CN_TOTAL_PRODUCTS = "\u5546\u54c1\u603b\u6570"
CN_AVG_WEIGHT = "\u5e73\u5747\u91cd\u91cf(pounds)"
CN_AVG_VOLUME = "\u5e73\u5747\u4f53\u79ef(in\u00b3)"
CN_AVG_MARGIN = "\u5e73\u5747\u6bdb\u5229\u7387"
CN_APLUS_SHARE = "A+\u5360\u6bd4"
CN_PRIMARY_COUNTRY = "\u5356\u5bb6\u6240\u5c5e\u5730"
CN_MONOPOLY_DEGREE = "\u5784\u65ad\u5ea6"
CN_NEW_PRODUCT_COUNT = "\u65b0\u54c1\u6570\u91cf"
CN_NEW_PRODUCT_SHARE = "\u65b0\u54c1\u5360\u6bd4"
CN_RETURN_RATE = "\u9000\u8d27\u7387"
CN_PEER_RETURN_RATE = "\u540c\u7c7b\u76ee\u9000\u8d27\u7387"
CN_SEARCH_PURCHASE_RATIO = "\u641c\u7d22\u8d2d\u4e70\u6bd4"
CN_PEER_SEARCH_PURCHASE_RATIO = "\u540c\u7c7b\u76ee\u641c\u7d22\u8d2d\u4e70\u6bd4"

GROUP_SAMPLE = "Listing\u6837\u672c\u6570"
GROUP_HEAD = "\u5934\u90e8Listing\u6570"
GROUP_NEW = "\u65b0\u54c1"
GROUP_ALL = "\u6240\u6709Listing"

SAMPLE_LABEL_PRODUCT = "\u5546\u54c1"
SAMPLE_LABEL_BRAND = "\u54c1\u724c"
SAMPLE_LABEL_SELLER = "\u5356\u5bb6"

FILENAME_PATTERN = re.compile(
    r"^Market-research(?:\((?P<scope>\d+)\))?(?P<body>.+)$",
    re.IGNORECASE,
)
BODY_PATTERN = re.compile(
    r"^(?P<query_seed>.+?)-(?P<marketplace>[A-Za-z]{2})(?:-(?P<time_window>.+))?$"
)
SELLER_SHARE_PATTERNS = {
    "FBA": re.compile(r"FBA:\s*(\d+(?:\.\d+)?)%", re.IGNORECASE),
    "AMZ": re.compile(r"AMZ:\s*(\d+(?:\.\d+)?)%", re.IGNORECASE),
    "FBM": re.compile(r"FBM:\s*(\d+(?:\.\d+)?)%", re.IGNORECASE),
}

NUMERIC_FIELDS = {
    "raw_row_number",
    "path_depth",
    "sample_products",
    "sample_brands",
    "sample_sellers",
    "monthly_sales_units",
    "monthly_sales_units_per_sample_listing",
    "monthly_revenue_usd",
    "avg_price_usd",
    "avg_review_count",
    "avg_star_rating",
    "avg_bsr",
    "avg_seller_count_per_listing",
    "total_products_in_market",
    "fba_share",
    "amz_share",
    "fbm_share",
    "seller_share_sum",
    "product_concentration",
    "brand_concentration",
    "seller_concentration",
    "top10_monthly_sales_units_per_listing",
    "top10_monthly_revenue_usd",
    "top10_avg_bsr",
    "top10_monopoly_degree",
    "new_product_count",
    "new_product_share",
    "new_products_monthly_sales_units_per_listing",
    "new_products_monthly_revenue_usd",
    "new_products_avg_price_usd",
    "new_products_avg_review_count",
    "new_products_avg_star_rating",
    "avg_weight_lb",
    "avg_volume_in3",
    "avg_gross_margin",
    "aplus_share",
    "return_rate",
    "peer_return_rate",
    "search_purchase_ratio",
    "peer_search_purchase_ratio",
    "primary_seller_country_share",
}


@dataclass
class SourceWorkbook:
    workbook_path: Path
    sheet_name: str
    notes_present: bool
    group_row: list[str]
    columns: list[dict[str, Any]]
    data_rows: list[list[Any]]


@dataclass
class RunPaths:
    run_id: str
    outputs_root: Path
    output_run_root: Path
    output_xlsx_dir: Path
    output_csv_dir: Path
    output_jsonl_dir: Path
    output_summary_dir: Path
    logs_root: Path
    log_run_root: Path
    run_log_path: Path
    warnings_path: Path
    errors_path: Path
    run_summary_path: Path
    archive_root: Path
    processed_root: Path
    processed_raw_inputs_dir: Path
    processed_manifest_dir: Path
    failed_root: Path
    failed_raw_inputs_dir: Path
    failed_manifest_dir: Path


@dataclass
class FileRecord:
    source_file: str
    status: str
    raw_input_root_before: str
    raw_input_archive_after: str = ""
    output_xlsx: str = ""
    output_csv: str = ""
    output_jsonl: str = ""
    log_file: str = ""
    warning_count: int = 0
    error_count: int = 0
    row_input: int = 0
    row_output: int = 0
    marketplace: str = ""
    batch_id: str = ""
    missing_source_columns: list[str] = field(default_factory=list)
    filename_issues: list[str] = field(default_factory=list)
    issue_counts: dict[str, int] = field(default_factory=dict)
    recognized_parent_paths: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


def now_local() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


def is_missing(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def as_text(value: Any) -> str:
    if is_missing(value):
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def slugify(value: str, separator: str = "_") -> str:
    text = re.sub(r"[^A-Za-z0-9]+", separator, value).strip(separator)
    text = re.sub(fr"{re.escape(separator)}+", separator, text)
    return text.lower() or "unknown"


def parse_args() -> argparse.Namespace:
    script_dir = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(
        description="Convert SellerSprite market export workbooks into M02 intermediate tables."
    )
    parser.add_argument("--input-dir", default=str(script_dir.parent))
    parser.add_argument("--output-dir", default=str(script_dir / "outputs"))
    parser.add_argument("--glob", default="Market-research*.xlsx")
    parser.add_argument("--overwrite", action="store_true")
    parser.add_argument("--debug", action="store_true")
    parser.add_argument(
        "--hard-delete-root-input-after-success",
        action="store_true",
        help="Copy successful inputs into archive and then delete the original root input after archive succeeds.",
    )
    return parser.parse_args()


def load_schema(schema_path: Path) -> dict[str, Any]:
    return json.loads(schema_path.read_text(encoding="utf-8"))


def normalize_group_row(values: list[Any]) -> list[str]:
    groups: list[str] = []
    current = ""
    for value in values:
        text = as_text(value)
        if text:
            current = text
        groups.append(current)
    return groups


def load_source_workbook(path: Path) -> SourceWorkbook:
    workbook = load_workbook(path, read_only=True, data_only=True)
    notes_present = "Notes" in workbook.sheetnames
    data_sheets = [sheet for sheet in workbook.worksheets if sheet.title != "Notes"]
    if not data_sheets:
        raise ValueError("No non-Notes worksheet was found.")

    sheet = data_sheets[0]
    group_row_raw = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    header_row_raw = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))
    group_row = normalize_group_row(list(group_row_raw))
    header_row = [as_text(value) for value in header_row_raw]

    columns: list[dict[str, Any]] = []
    for index, (group_name, header_name) in enumerate(zip(group_row, header_row)):
        columns.append(
            {
                "index": index,
                "group": group_name,
                "header": header_name,
                "group_norm": group_name.casefold(),
                "header_norm": header_name.casefold(),
            }
        )

    data_rows: list[list[Any]] = []
    for row in sheet.iter_rows(min_row=3, values_only=True):
        row_values = list(row)
        if any(not is_missing(value) for value in row_values):
            data_rows.append(row_values)

    return SourceWorkbook(
        workbook_path=path,
        sheet_name=sheet.title,
        notes_present=notes_present,
        group_row=group_row,
        columns=columns,
        data_rows=data_rows,
    )


def find_column_index(
    columns: list[dict[str, Any]],
    header_name: str,
    group_hint: str | None = None,
) -> int | None:
    header_norm = header_name.casefold()
    matches = [
        column["index"]
        for column in columns
        if column["header_norm"] == header_norm
        and (group_hint is None or group_hint.casefold() in column["group_norm"])
    ]
    return matches[0] if matches else None


def build_column_map(columns: list[dict[str, Any]]) -> tuple[dict[str, int | None], list[str]]:
    mapping = {
        "niche_en": find_column_index(columns, CN_NICHE_EN),
        "niche_zh": find_column_index(columns, CN_NICHE_ZH),
        "market_path_raw": find_column_index(columns, CN_MARKET_PATH),
        "sample_quantity_raw": find_column_index(columns, CN_SAMPLE_QUANTITY),
        "monthly_sales_units": find_column_index(columns, CN_MONTHLY_TOTAL_SALES),
        "monthly_sales_units_per_sample_listing": find_column_index(columns, CN_MONTHLY_AVG_SALES, GROUP_SAMPLE),
        "monthly_revenue_usd": find_column_index(columns, CN_MONTHLY_AVG_REVENUE, GROUP_SAMPLE),
        "avg_price_usd": find_column_index(columns, CN_AVG_PRICE, GROUP_SAMPLE),
        "avg_review_count": find_column_index(columns, CN_AVG_REVIEW_COUNT, GROUP_SAMPLE),
        "avg_star_rating": find_column_index(columns, CN_AVG_STAR, GROUP_SAMPLE),
        "avg_bsr": find_column_index(columns, CN_AVG_BSR, GROUP_SAMPLE),
        "avg_seller_count_per_listing": find_column_index(columns, CN_AVG_SELLER_COUNT),
        "seller_type_raw": find_column_index(columns, CN_SELLER_TYPE),
        "product_concentration": find_column_index(columns, CN_PRODUCT_CONCENTRATION),
        "brand_concentration": find_column_index(columns, CN_BRAND_CONCENTRATION),
        "seller_concentration": find_column_index(columns, CN_SELLER_CONCENTRATION),
        "total_products_in_market": find_column_index(columns, CN_TOTAL_PRODUCTS),
        "avg_weight_lb": find_column_index(columns, CN_AVG_WEIGHT),
        "avg_volume_in3": find_column_index(columns, CN_AVG_VOLUME),
        "avg_gross_margin": find_column_index(columns, CN_AVG_MARGIN),
        "aplus_share": find_column_index(columns, CN_APLUS_SHARE),
        "primary_seller_country_raw": find_column_index(columns, CN_PRIMARY_COUNTRY),
        "top10_monthly_sales_units_per_listing": find_column_index(columns, CN_MONTHLY_AVG_SALES, GROUP_HEAD),
        "top10_monopoly_degree": find_column_index(columns, CN_MONOPOLY_DEGREE),
        "top10_monthly_revenue_usd": find_column_index(columns, CN_MONTHLY_AVG_REVENUE, GROUP_HEAD),
        "top10_avg_bsr": find_column_index(columns, CN_AVG_BSR, GROUP_HEAD),
        "new_product_count": find_column_index(columns, CN_NEW_PRODUCT_COUNT),
        "new_product_share": find_column_index(columns, CN_NEW_PRODUCT_SHARE),
        "new_products_monthly_sales_units_per_listing": find_column_index(columns, CN_MONTHLY_AVG_SALES, GROUP_NEW),
        "new_products_monthly_revenue_usd": find_column_index(columns, CN_MONTHLY_AVG_REVENUE, GROUP_NEW),
        "new_products_avg_price_usd": find_column_index(columns, CN_AVG_PRICE, GROUP_NEW),
        "new_products_avg_review_count": find_column_index(columns, CN_AVG_REVIEW_COUNT, GROUP_NEW),
        "new_products_avg_star_rating": find_column_index(columns, CN_AVG_STAR, GROUP_NEW),
        "return_rate": find_column_index(columns, CN_RETURN_RATE),
        "peer_return_rate": find_column_index(columns, CN_PEER_RETURN_RATE),
        "search_purchase_ratio": find_column_index(columns, CN_SEARCH_PURCHASE_RATIO),
        "peer_search_purchase_ratio": find_column_index(columns, CN_PEER_SEARCH_PURCHASE_RATIO),
    }
    missing = [field_name for field_name, index in mapping.items() if index is None]
    return mapping, missing


def get_value(row: list[Any], column_map: dict[str, int | None], field_name: str) -> Any:
    index = column_map.get(field_name)
    if index is None or index >= len(row):
        return None
    return row[index]


def parse_filename_metadata(path: Path) -> tuple[dict[str, str], list[str]]:
    metadata = {
        "batch_id": slugify(path.stem),
        "query_seed": "",
        "marketplace": "",
        "time_window_raw": "",
    }
    issues: list[str] = []
    outer_match = FILENAME_PATTERN.match(path.stem)
    if not outer_match:
        issues.append("filename_pattern_unmatched")
        return metadata, issues

    body_match = BODY_PATTERN.match(outer_match.group("body") or "")
    if not body_match:
        issues.append("filename_body_unmatched")
        return metadata, issues

    metadata["query_seed"] = body_match.group("query_seed") or ""
    metadata["marketplace"] = (body_match.group("marketplace") or "").upper()
    metadata["time_window_raw"] = body_match.group("time_window") or ""
    return metadata, issues


def parse_number(value: Any) -> float | None:
    if is_missing(value):
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    normalized = (
        as_text(value)
        .replace(",", "")
        .replace("$", "")
        .replace("\uffe5", "")
        .replace("\u00a5", "")
        .replace("%", "")
        .replace("\n", " ")
    )
    match = re.search(r"-?\d+(?:\.\d+)?", normalized)
    return float(match.group(0)) if match else None


def parse_ratio(value: Any) -> float | None:
    if is_missing(value):
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        number = float(value)
        if 0 <= number <= 1:
            return number
        if 1 < number <= 100:
            return number / 100
        return number
    text = as_text(value)
    number = parse_number(text)
    if number is None:
        return None
    if "%" in text:
        return number / 100
    if 0 <= number <= 1:
        return number
    if 1 < number <= 100:
        return number / 100
    return number


def parse_numeric_field(value: Any, field_name: str, issues: list[str]) -> float | None:
    parsed = parse_number(value)
    if parsed is None and not is_missing(value):
        issues.append(f"invalid_{field_name}")
    return parsed


def parse_ratio_field(value: Any, field_name: str, issues: list[str]) -> float | None:
    parsed = parse_ratio(value)
    if parsed is None and not is_missing(value):
        issues.append(f"invalid_{field_name}")
    return parsed


def parse_labeled_block(
    value: Any,
    labels: dict[str, str],
    issue_prefix: str,
) -> tuple[dict[str, float | None], list[str]]:
    results = {target_field: None for target_field in labels.values()}
    issues: list[str] = []
    text = as_text(value)
    if not text:
        return results, issues

    unmatched_lines = 0
    for line in [line.strip() for line in text.split("\n") if line.strip()]:
        matched = False
        for label, target_field in labels.items():
            if label in line:
                parsed = parse_number(line)
                if parsed is None:
                    issues.append(f"invalid_{target_field}")
                else:
                    results[target_field] = parsed
                matched = True
                break
        if not matched:
            unmatched_lines += 1

    if unmatched_lines:
        issues.append(f"{issue_prefix}_unmatched_lines")
    return results, issues


def parse_seller_type_shares(value: Any) -> tuple[dict[str, float | None], float | None, str, list[str]]:
    results = {"fba_share": None, "amz_share": None, "fbm_share": None}
    issues: list[str] = []
    text = as_text(value)
    if not text:
        return results, None, "OK", issues

    matched_any = False
    invalid_parse = False
    for label, target_field in {"FBA": "fba_share", "AMZ": "amz_share", "FBM": "fbm_share"}.items():
        if re.search(rf"\b{label}\b", text, re.IGNORECASE):
            match = SELLER_SHARE_PATTERNS[label].search(text)
            if match:
                results[target_field] = float(match.group(1)) / 100.0
                matched_any = True
            else:
                invalid_parse = True
                issues.append(f"invalid_{target_field}")

    if not matched_any:
        invalid_parse = True

    parsed_values = [candidate for candidate in results.values() if candidate is not None]
    seller_share_sum = round(sum(parsed_values), 6) if parsed_values else None
    if invalid_parse:
        issues.append("invalid_seller_share_parse")
        return results, seller_share_sum, "INVALID_PARSE", issues
    if seller_share_sum is not None and seller_share_sum > 1.05:
        issues.append("invalid_seller_share_sum")
        return results, seller_share_sum, "INVALID_SUM", issues
    return results, seller_share_sum, "OK", issues


def parse_country_share(value: Any) -> tuple[str, float | None, list[str]]:
    text = as_text(value)
    if not text:
        return "", None, []
    parts = [part.strip() for part in text.split("\n") if part.strip()]
    country = ""
    share: float | None = None
    issues: list[str] = []
    for part in parts:
        if "%" in part or re.search(r"\d", part):
            candidate = parse_ratio(part)
            if candidate is None:
                issues.append("invalid_primary_seller_country_share")
            else:
                share = candidate
        elif not country:
            country = part
    if not country:
        issues.append("missing_primary_seller_country")
    return country, share, issues


def split_market_path(raw_path: str) -> dict[str, Any]:
    parts = [part.strip() for part in raw_path.split(":") if part and part.strip()]
    dept_l1 = parts[0] if parts else ""
    intermediates = parts[1:-1] if len(parts) > 1 else []
    parent_slots = ["", "", "", ""]
    for idx, value in enumerate(intermediates[:4]):
        parent_slots[idx] = value
    return {
        "path_depth": len(parts),
        "dept_l1": dept_l1,
        "parent_l2": parent_slots[0],
        "parent_l3": parent_slots[1],
        "parent_l4": parent_slots[2],
        "parent_l5": parent_slots[3],
        "niche_leaf": parts[-1] if parts else "",
        "path_key": " > ".join(parts),
    }


def normalize_row_types(row: dict[str, Any]) -> dict[str, Any]:
    for field_name, value in list(row.items()):
        if field_name in NUMERIC_FIELDS:
            if value is None or value == "":
                row[field_name] = None
            elif isinstance(value, float) and math.isnan(value):
                row[field_name] = None
            elif field_name in {"raw_row_number", "path_depth"} and value is not None:
                row[field_name] = int(value)
        else:
            row[field_name] = as_text(value)
    return row


def build_unique_run_id(paths_to_check: list[Path]) -> str:
    while True:
        candidate = datetime.now().strftime("%Y%m%d_%H%M%S")
        if not any((path / candidate).exists() for path in paths_to_check):
            return candidate
        time.sleep(1)


def build_run_paths(script_dir: Path, output_dir: Path) -> RunPaths:
    logs_root = script_dir / "logs"
    archive_root = script_dir / "archive"
    processed_root = archive_root / "processed"
    failed_root = archive_root / "failed"
    run_id = build_unique_run_id([output_dir, logs_root, processed_root, failed_root])
    output_run_root = output_dir / run_id
    log_run_root = logs_root / run_id
    processed_run_root = processed_root / run_id
    failed_run_root = failed_root / run_id
    run_paths = RunPaths(
        run_id=run_id,
        outputs_root=output_dir,
        output_run_root=output_run_root,
        output_xlsx_dir=output_run_root / "xlsx",
        output_csv_dir=output_run_root / "csv",
        output_jsonl_dir=output_run_root / "jsonl",
        output_summary_dir=output_run_root / "summaries",
        logs_root=logs_root,
        log_run_root=log_run_root,
        run_log_path=log_run_root / "run.log",
        warnings_path=log_run_root / "warnings.json",
        errors_path=log_run_root / "errors.json",
        run_summary_path=log_run_root / f"run_summary__{run_id}.json",
        archive_root=archive_root,
        processed_root=processed_run_root,
        processed_raw_inputs_dir=processed_run_root / "raw_inputs",
        processed_manifest_dir=processed_run_root / "manifests",
        failed_root=failed_run_root,
        failed_raw_inputs_dir=failed_run_root / "raw_inputs",
        failed_manifest_dir=failed_run_root / "manifests",
    )
    for path in [
        run_paths.output_xlsx_dir,
        run_paths.output_csv_dir,
        run_paths.output_jsonl_dir,
        run_paths.output_summary_dir,
        run_paths.log_run_root,
        run_paths.processed_raw_inputs_dir,
        run_paths.processed_manifest_dir,
        run_paths.failed_raw_inputs_dir,
        run_paths.failed_manifest_dir,
    ]:
        path.mkdir(parents=True, exist_ok=False)
    return run_paths


def scan_input_files(input_dir: Path, pattern: str, script_dir: Path) -> list[Path]:
    files: list[Path] = []
    for path in sorted(input_dir.glob(pattern)):
        if not path.is_file():
            continue
        try:
            path.relative_to(script_dir)
            continue
        except ValueError:
            pass
        files.append(path)
    return files


def make_market_cleaned_rows(
    source: SourceWorkbook,
    schema: dict[str, Any],
) -> tuple[pd.DataFrame, dict[str, Any]]:
    field_order = [field["field_name"] for field in schema["fields"]]
    column_map, missing_source_columns = build_column_map(source.columns)
    filename_meta, filename_issues = parse_filename_metadata(source.workbook_path)
    sample_scope_raw = next((group for group in source.group_row if GROUP_SAMPLE in group), "")
    head_scope_raw = next((group for group in source.group_row if GROUP_HEAD in group), "")
    new_product_window_raw = next((group for group in source.group_row if GROUP_NEW in group), "")

    rows: list[dict[str, Any]] = []
    issue_counter: Counter[str] = Counter()

    for offset, raw_row in enumerate(source.data_rows, start=3):
        issues: list[str] = []
        sample_counts, sample_count_issues = parse_labeled_block(
            get_value(raw_row, column_map, "sample_quantity_raw"),
            {
                SAMPLE_LABEL_PRODUCT: "sample_products",
                SAMPLE_LABEL_BRAND: "sample_brands",
                SAMPLE_LABEL_SELLER: "sample_sellers",
            },
            "sample_counts",
        )
        seller_shares, seller_share_sum, seller_share_parse_flag, seller_share_issues = parse_seller_type_shares(
            get_value(raw_row, column_map, "seller_type_raw")
        )
        country_text, country_share, country_issues = parse_country_share(
            get_value(raw_row, column_map, "primary_seller_country_raw")
        )

        market_path_raw = as_text(get_value(raw_row, column_map, "market_path_raw"))
        path_parts = split_market_path(market_path_raw) if market_path_raw else split_market_path("")
        if not market_path_raw:
            issues.append("missing_market_path")

        row = {
            "batch_id": filename_meta["batch_id"],
            "source_file": source.workbook_path.name,
            "source_sheet": source.sheet_name,
            "source_tool": SOURCE_TOOL,
            "route_type": ROUTE_TYPE,
            "stage_code": STAGE_CODE,
            "query_seed": filename_meta["query_seed"],
            "marketplace": filename_meta["marketplace"],
            "time_window_raw": filename_meta["time_window_raw"],
            "new_product_window_raw": new_product_window_raw,
            "sample_scope_raw": sample_scope_raw,
            "head_scope_raw": head_scope_raw,
            "raw_row_number": offset,
            "niche_en": as_text(get_value(raw_row, column_map, "niche_en")),
            "niche_zh": as_text(get_value(raw_row, column_map, "niche_zh")),
            "market_path_raw": market_path_raw,
            "path_depth": path_parts["path_depth"],
            "dept_l1": path_parts["dept_l1"],
            "parent_l2": path_parts["parent_l2"],
            "parent_l3": path_parts["parent_l3"],
            "parent_l4": path_parts["parent_l4"],
            "parent_l5": path_parts["parent_l5"],
            "niche_leaf": path_parts["niche_leaf"],
            "path_key": path_parts["path_key"],
            "sample_products": sample_counts["sample_products"],
            "sample_brands": sample_counts["sample_brands"],
            "sample_sellers": sample_counts["sample_sellers"],
            "monthly_sales_units": parse_numeric_field(get_value(raw_row, column_map, "monthly_sales_units"), "monthly_sales_units", issues),
            "monthly_sales_units_per_sample_listing": parse_numeric_field(get_value(raw_row, column_map, "monthly_sales_units_per_sample_listing"), "monthly_sales_units_per_sample_listing", issues),
            "monthly_revenue_usd": parse_numeric_field(get_value(raw_row, column_map, "monthly_revenue_usd"), "monthly_revenue_usd", issues),
            "avg_price_usd": parse_numeric_field(get_value(raw_row, column_map, "avg_price_usd"), "avg_price_usd", issues),
            "avg_review_count": parse_numeric_field(get_value(raw_row, column_map, "avg_review_count"), "avg_review_count", issues),
            "avg_star_rating": parse_numeric_field(get_value(raw_row, column_map, "avg_star_rating"), "avg_star_rating", issues),
            "avg_bsr": parse_numeric_field(get_value(raw_row, column_map, "avg_bsr"), "avg_bsr", issues),
            "avg_seller_count_per_listing": parse_numeric_field(get_value(raw_row, column_map, "avg_seller_count_per_listing"), "avg_seller_count_per_listing", issues),
            "total_products_in_market": parse_numeric_field(get_value(raw_row, column_map, "total_products_in_market"), "total_products_in_market", issues),
            "seller_type_raw": as_text(get_value(raw_row, column_map, "seller_type_raw")),
            "fba_share": seller_shares["fba_share"],
            "amz_share": seller_shares["amz_share"],
            "fbm_share": seller_shares["fbm_share"],
            "seller_share_sum": seller_share_sum,
            "seller_share_parse_flag": seller_share_parse_flag,
            "product_concentration": parse_ratio_field(get_value(raw_row, column_map, "product_concentration"), "product_concentration", issues),
            "brand_concentration": parse_ratio_field(get_value(raw_row, column_map, "brand_concentration"), "brand_concentration", issues),
            "seller_concentration": parse_ratio_field(get_value(raw_row, column_map, "seller_concentration"), "seller_concentration", issues),
            "top10_monthly_sales_units_per_listing": parse_numeric_field(get_value(raw_row, column_map, "top10_monthly_sales_units_per_listing"), "top10_monthly_sales_units_per_listing", issues),
            "top10_monthly_revenue_usd": parse_numeric_field(get_value(raw_row, column_map, "top10_monthly_revenue_usd"), "top10_monthly_revenue_usd", issues),
            "top10_avg_bsr": parse_numeric_field(get_value(raw_row, column_map, "top10_avg_bsr"), "top10_avg_bsr", issues),
            "top10_monopoly_degree": parse_numeric_field(get_value(raw_row, column_map, "top10_monopoly_degree"), "top10_monopoly_degree", issues),
            "new_product_count": parse_numeric_field(get_value(raw_row, column_map, "new_product_count"), "new_product_count", issues),
            "new_product_share": parse_ratio_field(get_value(raw_row, column_map, "new_product_share"), "new_product_share", issues),
            "new_products_monthly_sales_units_per_listing": parse_numeric_field(get_value(raw_row, column_map, "new_products_monthly_sales_units_per_listing"), "new_products_monthly_sales_units_per_listing", issues),
            "new_products_monthly_revenue_usd": parse_numeric_field(get_value(raw_row, column_map, "new_products_monthly_revenue_usd"), "new_products_monthly_revenue_usd", issues),
            "new_products_avg_price_usd": parse_numeric_field(get_value(raw_row, column_map, "new_products_avg_price_usd"), "new_products_avg_price_usd", issues),
            "new_products_avg_review_count": parse_numeric_field(get_value(raw_row, column_map, "new_products_avg_review_count"), "new_products_avg_review_count", issues),
            "new_products_avg_star_rating": parse_numeric_field(get_value(raw_row, column_map, "new_products_avg_star_rating"), "new_products_avg_star_rating", issues),
            "avg_weight_lb": parse_numeric_field(get_value(raw_row, column_map, "avg_weight_lb"), "avg_weight_lb", issues),
            "avg_volume_in3": parse_numeric_field(get_value(raw_row, column_map, "avg_volume_in3"), "avg_volume_in3", issues),
            "avg_gross_margin": parse_ratio_field(get_value(raw_row, column_map, "avg_gross_margin"), "avg_gross_margin", issues),
            "aplus_share": parse_ratio_field(get_value(raw_row, column_map, "aplus_share"), "aplus_share", issues),
            "return_rate": parse_ratio_field(get_value(raw_row, column_map, "return_rate"), "return_rate", issues),
            "peer_return_rate": parse_ratio_field(get_value(raw_row, column_map, "peer_return_rate"), "peer_return_rate", issues),
            "search_purchase_ratio": parse_numeric_field(get_value(raw_row, column_map, "search_purchase_ratio"), "search_purchase_ratio", issues),
            "peer_search_purchase_ratio": parse_numeric_field(get_value(raw_row, column_map, "peer_search_purchase_ratio"), "peer_search_purchase_ratio", issues),
            "primary_seller_country": country_text,
            "primary_seller_country_share": country_share,
            "keep_flag": "REVIEW_PENDING",
            "drop_reason": "",
            "next_action": "BUILD_NICHE_SHORTLIST",
            "next_object_type": "niche",
            "data_quality_flag": "OK",
            "parse_notes": "",
        }

        issues.extend(sample_count_issues)
        issues.extend(seller_share_issues)
        issues.extend(country_issues)
        issues.extend(filename_issues)
        if not row["query_seed"]:
            issues.append("missing_query_seed")
        if not row["marketplace"]:
            issues.append("missing_marketplace")
        if not row["time_window_raw"]:
            issues.append("missing_time_window_raw")
        if not row["niche_en"]:
            issues.append("missing_niche_en")

        deduped_issues = sorted(set(issue for issue in issues if issue))
        row["data_quality_flag"] = "WARN" if deduped_issues else "OK"
        row["parse_notes"] = "; ".join(deduped_issues)
        normalize_row_types(row)
        rows.append({field_name: row.get(field_name) for field_name in field_order})
        issue_counter.update(deduped_issues)

    diagnostics = {
        "missing_source_columns": missing_source_columns,
        "filename_issues": filename_issues,
        "issue_counts": dict(issue_counter),
        "notes_present": source.notes_present,
        "input_rows": len(source.data_rows),
    }
    return pd.DataFrame(rows, columns=field_order), diagnostics


def build_path_summary(frame: pd.DataFrame) -> pd.DataFrame:
    working = frame.copy()
    for column_name in ["dept_l1", "parent_l2", "parent_l3", "path_key", "niche_leaf"]:
        working[column_name] = working[column_name].fillna("")
    summary_rows: list[dict[str, Any]] = []

    def summarize(group_cols: list[str], summary_level: str) -> None:
        grouped = working.groupby(group_cols, dropna=False, sort=True)
        for group_values, group_frame in grouped:
            if not isinstance(group_values, tuple):
                group_values = (group_values,)
            value_map = dict(zip(group_cols, group_values))
            dept_l1 = value_map.get("dept_l1", "")
            parent_l2 = value_map.get("parent_l2", "")
            parent_l3 = value_map.get("parent_l3", "")
            if summary_level == "dept_l1":
                path_depth = 1 if dept_l1 else 0
                path_key = dept_l1
            elif summary_level == "dept_l1_parent_l2":
                path_depth = 2 if parent_l2 else (1 if dept_l1 else 0)
                path_key = " > ".join(part for part in [dept_l1, parent_l2] if part)
            else:
                path_depth = int(group_frame["path_depth"].dropna().iloc[0]) if not group_frame["path_depth"].dropna().empty else 0
                path_key = value_map.get("path_key", "")
                parent_l3 = group_frame["parent_l3"].fillna("").iloc[0]
            summary_rows.append(
                {
                    "summary_level": summary_level,
                    "path_key": path_key,
                    "dept_l1": dept_l1,
                    "parent_l2": parent_l2,
                    "parent_l3": parent_l3,
                    "path_depth": path_depth,
                    "niche_count": int(group_frame["niche_leaf"].replace("", pd.NA).nunique(dropna=True)),
                    "row_count": int(len(group_frame)),
                    "total_monthly_sales_units": group_frame["monthly_sales_units"].sum(min_count=1),
                    "avg_product_concentration": group_frame["product_concentration"].mean(),
                    "avg_brand_concentration": group_frame["brand_concentration"].mean(),
                    "avg_seller_concentration": group_frame["seller_concentration"].mean(),
                    "avg_new_product_share": group_frame["new_product_share"].mean(),
                    "avg_price_usd": group_frame["avg_price_usd"].mean(),
                }
            )

    summarize(["dept_l1"], "dept_l1")
    summarize(["dept_l1", "parent_l2"], "dept_l1_parent_l2")
    summarize(["path_key"], "full_path")
    return pd.DataFrame(summary_rows).sort_values(
        by=["summary_level", "total_monthly_sales_units", "path_key"],
        ascending=[True, False, True],
        kind="stable",
    ).reset_index(drop=True)


def build_run_log_frame(
    source: SourceWorkbook,
    diagnostics: dict[str, Any],
    output_workbook: Path,
    output_csv: Path,
    output_jsonl: Path,
    output_rows: int,
) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "status": "SUCCESS",
                "processed_at": now_local(),
                "input_file": source.workbook_path.name,
                "input_sheet": source.sheet_name,
                "notes_sheet_present": source.notes_present,
                "input_row_count": diagnostics["input_rows"],
                "output_row_count": output_rows,
                "output_workbook": str(output_workbook),
                "output_csv": str(output_csv),
                "output_jsonl": str(output_jsonl),
                "missing_source_column_count": len(diagnostics["missing_source_columns"]),
                "missing_source_columns": json.dumps(diagnostics["missing_source_columns"], ensure_ascii=False),
                "filename_issue_count": len(diagnostics["filename_issues"]),
                "filename_issues": json.dumps(diagnostics["filename_issues"], ensure_ascii=False),
                "row_warning_count": int(sum(diagnostics["issue_counts"].values())),
                "warning_codes": json.dumps(diagnostics["issue_counts"], ensure_ascii=False),
                "exception_count": 0,
                "exception_message": "",
            }
        ]
    )


def write_outputs(
    *,
    frame: pd.DataFrame,
    path_summary: pd.DataFrame,
    field_dictionary: pd.DataFrame,
    run_log: pd.DataFrame,
    xlsx_path: Path,
    csv_path: Path,
    jsonl_path: Path,
    overwrite: bool,
) -> None:
    for output_path in [xlsx_path, csv_path, jsonl_path]:
        if output_path.exists() and not overwrite:
            raise FileExistsError(f"Output already exists and --overwrite was not set: {output_path}")
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    jsonl_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        frame.to_excel(writer, sheet_name="market_cleaned", index=False)
        path_summary.to_excel(writer, sheet_name="path_summary", index=False)
        field_dictionary.to_excel(writer, sheet_name="field_dictionary", index=False)
        run_log.to_excel(writer, sheet_name="run_log", index=False)
    frame.to_csv(csv_path, index=False, encoding="utf-8-sig")
    frame.to_json(jsonl_path, orient="records", lines=True, force_ascii=False)


def transform_file_to_outputs(
    *,
    path: Path,
    run_paths: RunPaths,
    schema: dict[str, Any],
    overwrite: bool,
) -> tuple[FileRecord, dict[str, Any]]:
    source = load_source_workbook(path)
    frame, diagnostics = make_market_cleaned_rows(source, schema)
    path_summary = build_path_summary(frame)
    field_dictionary = pd.DataFrame(schema["fields"])

    marketplace_series = frame["marketplace"].fillna("").astype(str)
    batch_series = frame["batch_id"].fillna("").astype(str)
    marketplace = marketplace_series.iloc[0] if not marketplace_series.empty and marketplace_series.iloc[0] else "UNKNOWN"
    batch_id = batch_series.iloc[0] if not batch_series.empty and batch_series.iloc[0] else slugify(path.stem)
    basename = f"M02_market_cleaned__{marketplace}__{batch_id}"

    xlsx_path = run_paths.output_xlsx_dir / f"{basename}.xlsx"
    csv_path = run_paths.output_csv_dir / f"{basename}.csv"
    jsonl_path = run_paths.output_jsonl_dir / f"{basename}.jsonl"
    run_log_frame = build_run_log_frame(source, diagnostics, xlsx_path, csv_path, jsonl_path, len(frame))

    write_outputs(
        frame=frame,
        path_summary=path_summary,
        field_dictionary=field_dictionary,
        run_log=run_log_frame,
        xlsx_path=xlsx_path,
        csv_path=csv_path,
        jsonl_path=jsonl_path,
        overwrite=overwrite,
    )

    warning_count = int(sum(diagnostics["issue_counts"].values()))
    record = FileRecord(
        source_file=path.name,
        status="success",
        raw_input_root_before=str(path),
        output_xlsx=str(xlsx_path),
        output_csv=str(csv_path),
        output_jsonl=str(jsonl_path),
        log_file=str(run_paths.run_log_path),
        warning_count=warning_count,
        error_count=0,
        row_input=diagnostics["input_rows"],
        row_output=int(len(frame)),
        marketplace=marketplace,
        batch_id=batch_id,
        missing_source_columns=diagnostics["missing_source_columns"],
        filename_issues=diagnostics["filename_issues"],
        issue_counts=diagnostics["issue_counts"],
        recognized_parent_paths=sorted(
            {
                value
                for value in frame["parent_l2"].dropna().astype(str).tolist()
                if value.strip()
            }
        ),
        errors=[],
    )
    return record, {
        "source": source,
        "diagnostics": diagnostics,
        "path_summary_rows": int(len(path_summary)),
    }


def archive_file(
    source_path: Path,
    archive_dir: Path,
    *,
    hard_delete_after_archive: bool,
) -> Path:
    archive_dir.mkdir(parents=True, exist_ok=True)
    destination = archive_dir / source_path.name
    if destination.exists():
        raise FileExistsError(f"Archive destination already exists: {destination}")

    if hard_delete_after_archive:
        shutil.copy2(source_path, destination)
        if not destination.exists():
            raise FileNotFoundError(f"Archive copy was not created: {destination}")
        source_path.unlink()
        return destination

    shutil.move(str(source_path), str(destination))
    return destination


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def append_log(log_lines: list[str], message: str) -> None:
    log_lines.append(f"[{now_local()}] {message}")


def build_manifest(
    *,
    run_paths: RunPaths,
    started_at: str,
    finished_at: str,
    input_dir: Path,
    matched_files: list[Path],
    file_records: list[FileRecord],
) -> dict[str, Any]:
    success_count = sum(1 for record in file_records if record.status == "success")
    failure_count = sum(1 for record in file_records if record.status == "failed")
    return {
        "run_id": run_paths.run_id,
        "started_at": started_at,
        "finished_at": finished_at,
        "input_dir": str(input_dir),
        "matched_files": [str(path) for path in matched_files],
        "success_count": success_count,
        "failure_count": failure_count,
        "output_root": str(run_paths.output_run_root),
        "archive_processed_root": str(run_paths.processed_root),
        "archive_failed_root": str(run_paths.failed_root),
        "files": [
            {
                "source_file": record.source_file,
                "status": record.status,
                "raw_input_root_before": record.raw_input_root_before,
                "raw_input_archive_after": record.raw_input_archive_after,
                "output_xlsx": record.output_xlsx,
                "output_csv": record.output_csv,
                "output_jsonl": record.output_jsonl,
                "log_file": record.log_file,
                "warning_count": record.warning_count,
                "error_count": record.error_count,
                "row_input": record.row_input,
                "row_output": record.row_output,
                "marketplace": record.marketplace,
                "batch_id": record.batch_id,
                "missing_source_columns": record.missing_source_columns,
                "filename_issues": record.filename_issues,
                "issue_counts": record.issue_counts,
                "recognized_parent_paths": record.recognized_parent_paths,
                "errors": record.errors,
            }
            for record in file_records
        ],
    }


def main() -> int:
    args = parse_args()
    script_dir = Path(__file__).resolve().parent
    input_dir = Path(args.input_dir).resolve()
    output_dir = Path(args.output_dir).resolve()
    schema = load_schema(script_dir / "schema" / "m02_market_cleaned_schema.json")

    input_files = scan_input_files(input_dir, args.glob, script_dir)
    if not input_files:
        print(f"[scan] no files matched {args.glob!r} in {input_dir}")
        return 1

    run_paths = build_run_paths(script_dir, output_dir)
    run_started_at = now_local()
    log_lines: list[str] = []
    append_log(log_lines, f"run_id={run_paths.run_id}")
    append_log(log_lines, f"input_dir={input_dir}")
    append_log(log_lines, f"matched_files={len(input_files)}")
    append_log(log_lines, f"output_root={run_paths.output_run_root}")

    warnings_records: list[dict[str, Any]] = []
    errors_records: list[dict[str, Any]] = []
    file_records: list[FileRecord] = []
    output_index_rows: list[dict[str, Any]] = []

    print(f"[scan] found {len(input_files)} file(s) in {input_dir}")
    print(f"[run_id] {run_paths.run_id}")
    print(f"[output_root] {run_paths.output_run_root}")

    for source_path in input_files:
        append_log(log_lines, f"start file={source_path.name}")
        file_record: FileRecord | None = None
        try:
            file_record, _ = transform_file_to_outputs(
                path=source_path,
                run_paths=run_paths,
                schema=schema,
                overwrite=args.overwrite,
            )
            archived_path = archive_file(
                source_path,
                run_paths.processed_raw_inputs_dir,
                hard_delete_after_archive=args.hard_delete_root_input_after_success,
            )
            file_record.raw_input_archive_after = str(archived_path)
            append_log(
                log_lines,
                f"success file={source_path.name} rows_in={file_record.row_input} "
                f"rows_out={file_record.row_output} archived_to={archived_path}",
            )
            print(
                f"[ok] {source_path.name}: input_rows={file_record.row_input} "
                f"output_rows={file_record.row_output} xlsx={file_record.output_xlsx}"
            )
            if file_record.warning_count:
                print(
                    f"[warn] {source_path.name}: warning_count={file_record.warning_count} "
                    f"issue_counts={json.dumps(file_record.issue_counts, ensure_ascii=False)}"
                )
        except Exception as exc:  # noqa: BLE001
            trace = traceback.format_exc()
            archive_after = ""
            archive_error = ""
            if source_path.exists():
                try:
                    failed_archive = archive_file(
                        source_path,
                        run_paths.failed_raw_inputs_dir,
                        hard_delete_after_archive=False,
                    )
                    archive_after = str(failed_archive)
                except Exception as archive_exc:  # noqa: BLE001
                    archive_error = str(archive_exc)
            file_record = file_record or FileRecord(
                source_file=source_path.name,
                status="failed",
                raw_input_root_before=str(source_path),
                log_file=str(run_paths.run_log_path),
            )
            file_record.status = "failed"
            file_record.raw_input_archive_after = archive_after
            file_record.error_count += 1
            file_record.errors.append(str(exc))
            if archive_error:
                file_record.error_count += 1
                file_record.errors.append(f"archive_failed: {archive_error}")
            errors_records.append(
                {
                    "source_file": source_path.name,
                    "errors": file_record.errors,
                    "traceback": trace,
                    "raw_input_archive_after": archive_after,
                }
            )
            append_log(log_lines, f"failed file={source_path.name} error={exc}")
            if archive_error:
                append_log(log_lines, f"failed file={source_path.name} archive_error={archive_error}")
            print(f"[fail] {source_path.name}: {exc}")
            if archive_after:
                print(f"[fail-archived] {source_path.name}: {archive_after}")
            if args.debug:
                traceback.print_exc()

        if file_record.warning_count or file_record.missing_source_columns or file_record.filename_issues:
            warnings_records.append(
                {
                    "source_file": file_record.source_file,
                    "warning_count": file_record.warning_count,
                    "missing_source_columns": file_record.missing_source_columns,
                    "filename_issues": file_record.filename_issues,
                    "issue_counts": file_record.issue_counts,
                    "recognized_parent_paths": file_record.recognized_parent_paths,
                }
            )

        file_records.append(file_record)
        output_index_rows.append(
            {
                "run_id": run_paths.run_id,
                "source_file": file_record.source_file,
                "marketplace": file_record.marketplace,
                "batch_id": file_record.batch_id,
                "output_xlsx": file_record.output_xlsx,
                "output_csv": file_record.output_csv,
                "output_jsonl": file_record.output_jsonl,
                "row_count": file_record.row_output,
                "status": file_record.status,
            }
        )

    run_finished_at = now_local()
    append_log(log_lines, f"finished success={sum(1 for r in file_records if r.status == 'success')} failed={sum(1 for r in file_records if r.status == 'failed')}")

    run_paths.run_log_path.write_text("\n".join(log_lines) + "\n", encoding="utf-8")
    write_json(run_paths.warnings_path, warnings_records)
    write_json(run_paths.errors_path, errors_records)

    output_index = pd.DataFrame(output_index_rows)
    output_index.to_csv(run_paths.output_summary_dir / "output_index.csv", index=False, encoding="utf-8-sig")

    manifest = build_manifest(
        run_paths=run_paths,
        started_at=run_started_at,
        finished_at=run_finished_at,
        input_dir=input_dir,
        matched_files=input_files,
        file_records=file_records,
    )
    write_json(run_paths.processed_manifest_dir / "run_manifest.json", manifest)
    if any(record.status == "failed" for record in file_records):
        write_json(run_paths.failed_manifest_dir / "run_manifest.json", manifest)

    run_summary = {
        "run_id": run_paths.run_id,
        "run_started_at": run_started_at,
        "run_finished_at": run_finished_at,
        "input_dir": str(input_dir),
        "output_root": str(run_paths.output_run_root),
        "log_root": str(run_paths.log_run_root),
        "archive_processed_root": str(run_paths.processed_root),
        "archive_failed_root": str(run_paths.failed_root),
        "processed_file_count": len(file_records),
        "success_count": sum(1 for record in file_records if record.status == "success"),
        "failure_count": sum(1 for record in file_records if record.status == "failed"),
        "matched_files": [str(path) for path in input_files],
        "results": [
            {
                "source_file": record.source_file,
                "status": record.status,
                "raw_input_archive_after": record.raw_input_archive_after,
                "output_xlsx": record.output_xlsx,
                "output_csv": record.output_csv,
                "output_jsonl": record.output_jsonl,
                "warning_count": record.warning_count,
                "error_count": record.error_count,
                "row_input": record.row_input,
                "row_output": record.row_output,
            }
            for record in file_records
        ],
    }
    write_json(run_paths.run_summary_path, run_summary)

    print(f"[log_root] {run_paths.log_run_root}")
    print(f"[warnings_json] {run_paths.warnings_path}")
    print(f"[errors_json] {run_paths.errors_path}")
    print(f"[output_index] {run_paths.output_summary_dir / 'output_index.csv'}")
    print(f"[processed_manifest] {run_paths.processed_manifest_dir / 'run_manifest.json'}")
    print(f"[processed_archive] {run_paths.processed_root}")
    print(f"[failed_archive] {run_paths.failed_root}")
    return 1 if any(record.status == "failed" for record in file_records) else 0


if __name__ == "__main__":
    sys.exit(main())
