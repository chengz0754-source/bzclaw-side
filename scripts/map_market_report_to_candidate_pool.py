from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DEFAULT_CANDIDATE_RELATIVE = Path("inputs/selection_run_current/03_候选市场与候选品初筛池.csv")
TEMPLATE_CANDIDATE_RELATIVE = Path("templates/selection_csv_cn_reference/03_候选市场与候选品初筛池.csv")
DEFAULT_MARKET_DIR_RELATIVE = Path("runs/manual/10_market")
OUTPUTS_ROOT_RELATIVE = Path("outputs/selection_runs")
RULE_BASED_PREFIX = "RULE_BASED_PRELIMINARY"
CSV_READ_ENCODINGS = ("utf-8-sig", "utf-8", "gb18030", "gbk")
CANONICAL_WORKBOOK_PREFIX = "market-report-"
DIAGNOSTIC_WORKBOOK_PREFIXES = ("diag-", "archive-")
WORKBOOK_SELECTION_RULE = (
    "prefer newest market-report-*.xlsx; fallback to newest non-diagnostic .xlsx; "
    "fail closed if only diagnostic/archive copies remain"
)

TARGET_FIELDS = [
    "运行名称",
    "种子关键词",
    "站点",
    "市场路径",
    "候选市场名称",
    "商品样本数",
    "品牌样本数",
    "卖家样本数",
    "月总销量",
    "月均销量",
    "月均销售额",
    "平均价格",
    "平均评分数",
    "平均星级",
    "新品数量",
    "新品占比",
    "商品集中度",
    "品牌集中度",
    "卖家集中度",
    "竞争强度判断",
    "是否进入候选池",
    "初筛结论",
    "初筛原因",
    "备注",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="把 SellerSprite 市场表映射为 03_候选市场与候选品初筛池.csv，并生成映射报告。"
    )
    parser.add_argument(
        "--market-workbook",
        default=None,
        help="市场表 xlsx 路径；未传时自动读取 runs/manual/10_market/ 中最新的 .xlsx。",
    )
    parser.add_argument(
        "--candidate-csv",
        default=str(DEFAULT_CANDIDATE_RELATIVE),
        help="候选池 CSV 路径，默认写入 inputs/selection_run_current/03_候选市场与候选品初筛池.csv。",
    )
    parser.add_argument("--run-name", default=None, help="覆盖写入的运行名称。")
    parser.add_argument("--site", default=None, help="覆盖写入的站点。")
    parser.add_argument("--seed-keyword", default=None, help="覆盖写入的种子关键词。")
    parser.add_argument("--market-path", default=None, help="覆盖写入的市场路径。")
    parser.add_argument(
        "--output-dir",
        default=None,
        help="映射产物目录；未传时默认生成到 outputs/selection_runs/<timestamp>/02_generated_outputs/。",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="只分析和生成报告，不落盘更新候选池 CSV。",
    )
    mode_group = parser.add_mutually_exclusive_group()
    mode_group.add_argument(
        "--append",
        action="store_true",
        help="总是把映射结果追加到候选池 CSV 尾部。",
    )
    mode_group.add_argument(
        "--replace-row",
        action="store_true",
        help="按候选市场名称 + 市场路径替换已有行；找不到则追加。默认行为。",
    )
    return parser.parse_args()


def ensure_within_repo(repo_root: Path, path: Path, label: str) -> Path:
    resolved = path.resolve()
    if not resolved.is_relative_to(repo_root.resolve()):
        raise ValueError(f"{label} 超出仓库根目录：{resolved}")
    return resolved


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\r", " ").replace("\n", " / ").strip()


def normalize_header(value: Any) -> str:
    text = normalize_text(value).lower()
    text = text.replace("（", "(").replace("）", ")")
    text = re.sub(r"[\s_/:\-]+", "", text)
    text = text.replace("$", "")
    text = text.replace("(", "").replace(")", "")
    return text


def safe_number_text(value: Any) -> str:
    if value is None or str(value).strip() == "":
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return f"{value:.4f}".rstrip("0").rstrip(".")
    text = str(value).strip()
    text = text.replace(",", "")
    return text


def format_ratio(value: Any) -> str:
    if value is None or str(value).strip() == "":
        return ""
    if isinstance(value, (int, float)):
        numeric = float(value)
        if numeric <= 1:
            text = f"{numeric * 100:.2f}".rstrip("0").rstrip(".")
            return f"{text}%"
        text = f"{numeric:.2f}".rstrip("0").rstrip(".")
        return f"{text}%"
    text = str(value).strip()
    if text.endswith("%"):
        return text
    try:
        numeric = float(text)
    except ValueError:
        return text
    if numeric <= 1:
        text = f"{numeric * 100:.2f}".rstrip("0").rstrip(".")
        return f"{text}%"
    text = f"{numeric:.2f}".rstrip("0").rstrip(".")
    return f"{text}%"


def parse_sample_counts(text: str) -> dict[str, str]:
    counts = {"商品样本数": "", "品牌样本数": "", "卖家样本数": ""}
    if not text:
        return counts
    for label, target_field in [("商品", "商品样本数"), ("品牌", "品牌样本数"), ("卖家", "卖家样本数")]:
        match = re.search(rf"{label}[：:]\s*(\d+)", text)
        if match:
            counts[target_field] = match.group(1)
    return counts


HEADER_ALIASES = {
    "细分市场": {"细分市场", "细分市场翻译", "市场名称", "细分类目"},
    "市场路径": {"市场路径", "类目路径", "市场类目路径"},
    "样本数量": {"样本数量", "样本数"},
    "月总销量": {"月总销量"},
    "月均销量": {"月均销量"},
    "月均销售额": {"月均销售额", "月均销售额$", "月均销售额usd"},
    "平均价格": {"平均价格", "平均价格$", "平均单价"},
    "平均评分数": {"平均评分数", "平均评论数", "平均review数"},
    "平均星级": {"平均星级", "平均评分"},
    "新品数量": {"新品数量"},
    "新品占比": {"新品占比", "新品率"},
    "商品集中度": {"商品集中度"},
    "品牌集中度": {"品牌集中度"},
    "卖家集中度": {"卖家集中度"},
}
HEADER_ALIASES = {
    field: {normalize_header(alias) for alias in aliases}
    for field, aliases in HEADER_ALIASES.items()
}


def score_header_row(row_values: list[Any]) -> tuple[int, dict[str, list[int]]]:
    matches: dict[str, list[int]] = defaultdict(list)
    score = 0
    for index, value in enumerate(row_values):
        normalized = normalize_header(value)
        if not normalized:
            continue
        for logical_field, aliases in HEADER_ALIASES.items():
            if normalized in aliases:
                matches[logical_field].append(index)
                score += 1
    return score, matches


def find_market_sheet_and_header(workbook_path: Path) -> tuple[str, int, list[str], dict[str, list[int]]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    best_sheet_name = ""
    best_header_row = 0
    best_headers: list[str] = []
    best_mapping: dict[str, list[int]] = {}
    best_score = -1

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        for row_index, row in enumerate(
            worksheet.iter_rows(min_row=1, max_row=10, values_only=True),
            start=1,
        ):
            row_values = list(row)
            score, mapping = score_header_row(row_values)
            if score > best_score:
                best_score = score
                best_sheet_name = sheet_name
                best_header_row = row_index
                best_headers = [normalize_text(value) for value in row_values]
                best_mapping = mapping

    workbook.close()
    if best_score < 5:
        raise RuntimeError("无法识别市场表的标题行，命中稳定字段过少。")
    return best_sheet_name, best_header_row, best_headers, best_mapping


def read_sheet_rows(workbook_path: Path, sheet_name: str, header_row_index: int) -> list[tuple[int, list[Any]]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name]
    rows: list[tuple[int, list[Any]]] = []
    for row_index, row in enumerate(
        worksheet.iter_rows(min_row=header_row_index + 1, values_only=True),
        start=header_row_index + 1,
    ):
        values = list(row)
        if not any(normalize_text(value) for value in values):
            continue
        rows.append((row_index, values))
    workbook.close()
    return rows


def first_value(row_values: list[Any], mapping: dict[str, list[int]], field: str, occurrence: int = 0) -> Any:
    indexes = mapping.get(field, [])
    if occurrence >= len(indexes):
        return None
    index = indexes[occurrence]
    if index >= len(row_values):
        return None
    return row_values[index]


def guess_site_from_workbook(path: Path) -> str:
    name = path.name.lower()
    if "us" in name:
        return "US"
    return ""


def load_csv_rows(path: Path) -> list[list[str]]:
    raw_bytes = path.read_bytes()
    decode_errors: list[str] = []
    for encoding in CSV_READ_ENCODINGS:
        try:
            return list(csv.reader(raw_bytes.decode(encoding).splitlines()))
        except UnicodeDecodeError as exc:
            decode_errors.append(f"{encoding}@{exc.start}:{exc.reason}")
    error_text = " | ".join(decode_errors)
    raise RuntimeError(f"无法读取 CSV 编码：{path}；已尝试 {CSV_READ_ENCODINGS}；错误：{error_text}")


def load_context_defaults(repo_root: Path) -> dict[str, str]:
    defaults = {"运行名称": "", "种子关键词": "", "站点": ""}
    sources = [
        repo_root / "inputs" / "selection_run_current" / "00_选品运行目标与边界.csv",
        repo_root / "inputs" / "selection_run_current" / "01_市场入口与筛选参数.csv",
    ]
    for path in sources:
        if not path.exists():
            continue
        rows = load_csv_rows(path)
        if len(rows) < 2:
            continue
        row_map = {header: rows[1][idx] if idx < len(rows[1]) else "" for idx, header in enumerate(rows[0])}
        if not defaults["运行名称"] and row_map.get("运行名称"):
            defaults["运行名称"] = row_map["运行名称"]
        if not defaults["种子关键词"]:
            if row_map.get("种子关键词"):
                defaults["种子关键词"] = row_map["种子关键词"]
            elif row_map.get("方向词"):
                defaults["种子关键词"] = row_map["方向词"]
        if not defaults["站点"] and row_map.get("站点"):
            defaults["站点"] = row_map["站点"]
    return defaults


def infer_competition_level(
    product_concentration: str,
    brand_concentration: str,
    seller_concentration: str,
    avg_review_count: str,
) -> str:
    values: list[float] = []
    for item in (product_concentration, brand_concentration, seller_concentration):
        try:
            values.append(float(item))
        except ValueError:
            continue
    try:
        review_count = float(avg_review_count)
    except ValueError:
        review_count = 0.0

    concentration_score = max(values) if values else 0.0
    if concentration_score >= 0.55 or review_count >= 1500:
        level = "高"
    elif concentration_score >= 0.4 or review_count >= 500:
        level = "中高"
    elif concentration_score >= 0.25 or review_count >= 150:
        level = "中"
    else:
        level = "中低"
    return f"{RULE_BASED_PREFIX}:{level}"


def build_mapped_row(
    row_index: int,
    row_values: list[Any],
    header_mapping: dict[str, list[int]],
    workbook_path: Path,
    sheet_name: str,
    args: argparse.Namespace,
    context_defaults: dict[str, str],
) -> tuple[dict[str, str], list[str], list[str]]:
    row: dict[str, str] = {field: "" for field in TARGET_FIELDS}
    mapped_fields: list[str] = []
    not_mapped_fields: list[str] = []

    sample_counts_text = normalize_text(first_value(row_values, header_mapping, "样本数量"))
    sample_counts = parse_sample_counts(sample_counts_text)
    market_name = normalize_text(first_value(row_values, header_mapping, "细分市场"))
    market_path = args.market_path or normalize_text(first_value(row_values, header_mapping, "市场路径"))
    monthly_sales = safe_number_text(first_value(row_values, header_mapping, "月总销量"))
    avg_monthly_sales = safe_number_text(first_value(row_values, header_mapping, "月均销量", 0))
    avg_monthly_revenue = safe_number_text(first_value(row_values, header_mapping, "月均销售额", 0))
    avg_price = safe_number_text(first_value(row_values, header_mapping, "平均价格", 0))
    avg_reviews = safe_number_text(first_value(row_values, header_mapping, "平均评分数", 0))
    avg_rating = safe_number_text(first_value(row_values, header_mapping, "平均星级", 0))
    new_product_count = safe_number_text(first_value(row_values, header_mapping, "新品数量"))
    new_product_ratio = format_ratio(first_value(row_values, header_mapping, "新品占比"))
    product_concentration = safe_number_text(first_value(row_values, header_mapping, "商品集中度"))
    brand_concentration = safe_number_text(first_value(row_values, header_mapping, "品牌集中度"))
    seller_concentration = safe_number_text(first_value(row_values, header_mapping, "卖家集中度"))

    row.update(
        {
            "运行名称": args.run_name or context_defaults.get("运行名称", ""),
            "种子关键词": args.seed_keyword or context_defaults.get("种子关键词", ""),
            "站点": args.site or context_defaults.get("站点", "") or guess_site_from_workbook(workbook_path),
            "市场路径": market_path,
            "候选市场名称": market_name,
            "商品样本数": sample_counts["商品样本数"],
            "品牌样本数": sample_counts["品牌样本数"],
            "卖家样本数": sample_counts["卖家样本数"],
            "月总销量": monthly_sales,
            "月均销量": avg_monthly_sales,
            "月均销售额": avg_monthly_revenue,
            "平均价格": avg_price,
            "平均评分数": avg_reviews,
            "平均星级": avg_rating,
            "新品数量": new_product_count,
            "新品占比": new_product_ratio,
            "商品集中度": product_concentration,
            "品牌集中度": brand_concentration,
            "卖家集中度": seller_concentration,
            "竞争强度判断": infer_competition_level(
                product_concentration,
                brand_concentration,
                seller_concentration,
                avg_reviews,
            ),
            "是否进入候选池": f"{RULE_BASED_PREFIX}:待人工判断",
            "初筛结论": f"{RULE_BASED_PREFIX}:待人工复核",
            "初筛原因": f"{RULE_BASED_PREFIX}:基于市场表自动映射基础指标，需结合00~02人工判断",
            "备注": f"源工作簿={workbook_path.name}; 源工作表={sheet_name}; 源数据行={row_index}",
        }
    )

    for field in TARGET_FIELDS:
        if row.get(field):
            mapped_fields.append(field)
        else:
            not_mapped_fields.append(field)
    return row, mapped_fields, not_mapped_fields


def write_csv_atomic(path: Path, headers: list[str], rows: list[list[str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = path.with_suffix(path.suffix + ".tmp")
    with temp_path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(headers)
        writer.writerows(rows)
    temp_path.replace(path)


def ensure_candidate_csv(repo_root: Path, candidate_csv_path: Path) -> None:
    if candidate_csv_path.exists():
        return
    template_path = repo_root / TEMPLATE_CANDIDATE_RELATIVE
    if not template_path.exists():
        raise FileNotFoundError(f"模板缺失：{template_path}")
    candidate_csv_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template_path, candidate_csv_path)


def merge_candidate_rows(
    repo_root: Path,
    candidate_csv_path: Path,
    mapped_rows: list[dict[str, str]],
    append_mode: bool,
) -> tuple[list[str], list[list[str]], list[dict[str, str]]]:
    ensure_candidate_csv(repo_root, candidate_csv_path)
    existing_rows = load_csv_rows(candidate_csv_path)
    if not existing_rows:
        raise RuntimeError(f"候选池 CSV 为空：{candidate_csv_path}")
    headers = existing_rows[0]
    data_rows = existing_rows[1:]
    template_rows = load_csv_rows(repo_root / TEMPLATE_CANDIDATE_RELATIVE)
    template_example_rows = template_rows[1:]

    if data_rows == template_example_rows:
        data_rows = []

    row_change_log: list[dict[str, str]] = []
    for mapped_row in mapped_rows:
        ordered_row = [mapped_row.get(header, "") for header in headers]
        target_name = mapped_row.get("候选市场名称", "")
        target_path = mapped_row.get("市场路径", "")
        if append_mode:
            data_rows.append(ordered_row)
            row_change_log.append(
                {
                    "action": "APPENDED",
                    "候选市场名称": target_name,
                    "市场路径": target_path,
                }
            )
            continue

        replaced = False
        for index, existing_row in enumerate(data_rows):
            existing_map = {
                header: existing_row[pos] if pos < len(existing_row) else ""
                for pos, header in enumerate(headers)
            }
            if (
                existing_map.get("候选市场名称", "") == target_name
                and existing_map.get("市场路径", "") == target_path
            ):
                data_rows[index] = ordered_row
                row_change_log.append(
                    {
                        "action": "REPLACED",
                        "候选市场名称": target_name,
                        "市场路径": target_path,
                        "row_index": str(index + 2),
                    }
                )
                replaced = True
                break

        if not replaced:
            data_rows.append(ordered_row)
            row_change_log.append(
                {
                    "action": "APPENDED_ON_REPLACE_MODE",
                    "候选市场名称": target_name,
                    "市场路径": target_path,
                }
            )

    return headers, data_rows, row_change_log


def latest_market_workbook(repo_root: Path) -> Path:
    market_dir = repo_root / DEFAULT_MARKET_DIR_RELATIVE
    if not market_dir.exists():
        raise FileNotFoundError(f"市场表目录不存在：{market_dir}")
    candidates = sorted(market_dir.glob("*.xlsx"), key=lambda item: item.stat().st_mtime, reverse=True)
    if not candidates:
        raise FileNotFoundError(f"市场表目录中未找到 .xlsx：{market_dir}")

    keep_candidates = [
        item
        for item in candidates
        if not item.name.lower().startswith(DIAGNOSTIC_WORKBOOK_PREFIXES)
    ]
    canonical_candidates = [
        item for item in keep_candidates if item.name.lower().startswith(CANONICAL_WORKBOOK_PREFIX)
    ]
    if canonical_candidates:
        return canonical_candidates[0]
    if keep_candidates:
        return keep_candidates[0]

    diagnostic_names = ", ".join(item.name for item in candidates)
    raise FileNotFoundError(
        "runs/manual/10_market/ 仅剩 diagnostic/archive workbook，"
        f"未找到 keep-set 原始市场表：{diagnostic_names}"
    )


def default_output_dir(repo_root: Path) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return repo_root / OUTPUTS_ROOT_RELATIVE / timestamp / "02_generated_outputs"


def write_cleaned_csv(output_dir: Path, mapped_rows: list[dict[str, str]]) -> Path:
    cleaned_path = output_dir / "market_cleaned.csv"
    headers = ["源工作簿", "源工作表", "源数据行"] + TARGET_FIELDS
    rows: list[list[str]] = []
    for row in mapped_rows:
        rows.append(
            [row.get("_source_workbook", ""), row.get("_source_sheet", ""), row.get("_source_row", "")]
            + [row.get(field, "") for field in TARGET_FIELDS]
        )
    write_csv_atomic(cleaned_path, headers, rows)
    return cleaned_path


def build_report_markdown(
    workbook_path: Path,
    sheet_name: str,
    header_row_index: int,
    raw_headers: list[str],
    row_reports: list[dict[str, Any]],
    not_mapped_fields_union: list[str],
    row_change_log: list[dict[str, str]],
    candidate_csv_path: Path,
    cleaned_csv_path: Path,
    report_json_path: Path,
    dry_run: bool,
    append_mode: bool,
) -> str:
    lines = [
        "# 市场表到03候选池映射报告",
        "",
        f"- 市场表：`{workbook_path}`",
        f"- 识别到的工作表：`{sheet_name}`",
        f"- 标题行：第 `{header_row_index}` 行",
        f"- 目标候选池 CSV：`{candidate_csv_path}`",
        f"- 运行模式：`{'APPEND' if append_mode else 'REPLACE_ROW'}`",
        f"- 是否 dry-run：`{'YES' if dry_run else 'NO'}`",
        f"- 市场表自动选择规则：`{WORKBOOK_SELECTION_RULE}`",
        f"- 清洗结果：`{cleaned_csv_path}`",
        f"- JSON 报告：`{report_json_path}`",
        "",
        "## 识别到的关键表头",
        "",
    ]
    lines.extend([f"- `{header}`" for header in raw_headers if header] or ["- 无"])
    lines.extend(["", "## 行级映射结果", ""])
    for row_report in row_reports:
        lines.extend(
            [
                f"### 源数据行 {row_report['source_row']}",
                "",
                f"- 候选市场名称：`{row_report['candidate_name']}`",
                f"- 市场路径：`{row_report['market_path']}`",
                f"- 成功映射字段数：`{len(row_report['mapped_fields'])}`",
                "",
                "已映射字段：",
            ]
        )
        lines.extend([f"- `{field}`" for field in row_report["mapped_fields"]] or ["- 无"])
        lines.extend(["", "未映射字段："])
        lines.extend([f"- `{field}` = `NOT_MAPPED`" for field in row_report["not_mapped_fields"]] or ["- 无"])
        lines.extend(
            [
                "",
                "规则初判字段：",
                f"- `竞争强度判断` = `{row_report['rule_based']['竞争强度判断']}`",
                f"- `是否进入候选池` = `{row_report['rule_based']['是否进入候选池']}`",
                f"- `初筛结论` = `{row_report['rule_based']['初筛结论']}`",
                f"- `初筛原因` = `{row_report['rule_based']['初筛原因']}`",
                "",
            ]
        )

    lines.extend(["## 未映射字段汇总", ""])
    lines.extend([f"- `{field}`" for field in not_mapped_fields_union] or ["- 无"])
    lines.extend(["", "## 候选池更新结果", ""])
    lines.extend([f"- 更新行数：`{len(row_change_log)}`", "- 变更明细："])
    for item in row_change_log:
        lines.append(
            f"- `{item['action']}` | 候选市场名称=`{item.get('候选市场名称', '')}` | 市场路径=`{item.get('市场路径', '')}`"
        )
    if dry_run:
        lines.append("- dry-run 模式下未实际写入候选池 CSV。")
    return "\n".join(lines) + "\n"


def main() -> int:
    args = parse_args()
    repo_root = Path(__file__).resolve().parents[1]

    try:
        market_workbook = (
            ensure_within_repo(repo_root, Path(args.market_workbook), "market_workbook")
            if args.market_workbook
            else latest_market_workbook(repo_root)
        )
        candidate_csv_path = ensure_within_repo(repo_root, Path(args.candidate_csv), "candidate_csv")
        output_dir = (
            ensure_within_repo(repo_root, Path(args.output_dir), "output_dir")
            if args.output_dir
            else default_output_dir(repo_root)
        )
    except (FileNotFoundError, ValueError) as exc:
        print(f"执行失败：{exc}", file=sys.stderr)
        return 1

    if market_workbook.suffix.lower() != ".xlsx":
        print(f"执行失败：市场表必须是 .xlsx：{market_workbook}", file=sys.stderr)
        return 1

    try:
        sheet_name, header_row_index, raw_headers, header_mapping = find_market_sheet_and_header(market_workbook)
        source_rows = read_sheet_rows(market_workbook, sheet_name, header_row_index)
    except Exception as exc:
        print(f"执行失败：读取市场表时出错：{exc}", file=sys.stderr)
        return 1

    context_defaults = load_context_defaults(repo_root)
    mapped_rows: list[dict[str, str]] = []
    row_reports: list[dict[str, Any]] = []
    not_mapped_fields_union: set[str] = set()

    for row_index, row_values in source_rows:
        candidate_name = normalize_text(first_value(row_values, header_mapping, "细分市场"))
        market_path = args.market_path or normalize_text(first_value(row_values, header_mapping, "市场路径"))
        if not candidate_name and not market_path:
            continue
        mapped_row, mapped_fields, not_mapped_fields = build_mapped_row(
            row_index=row_index,
            row_values=row_values,
            header_mapping=header_mapping,
            workbook_path=market_workbook,
            sheet_name=sheet_name,
            args=args,
            context_defaults=context_defaults,
        )
        mapped_row["_source_workbook"] = market_workbook.name
        mapped_row["_source_sheet"] = sheet_name
        mapped_row["_source_row"] = str(row_index)
        mapped_rows.append(mapped_row)
        not_mapped_fields_union.update(not_mapped_fields)
        row_reports.append(
            {
                "source_row": row_index,
                "candidate_name": mapped_row["候选市场名称"],
                "market_path": mapped_row["市场路径"],
                "mapped_fields": mapped_fields,
                "not_mapped_fields": not_mapped_fields,
                "rule_based": {
                    "竞争强度判断": mapped_row["竞争强度判断"],
                    "是否进入候选池": mapped_row["是否进入候选池"],
                    "初筛结论": mapped_row["初筛结论"],
                    "初筛原因": mapped_row["初筛原因"],
                },
            }
        )

    if not mapped_rows:
        print("执行失败：未在市场表中识别到可映射的数据行。", file=sys.stderr)
        return 1

    output_dir.mkdir(parents=True, exist_ok=True)
    cleaned_csv_path = write_cleaned_csv(output_dir, mapped_rows)

    if candidate_csv_path.name == "03_候选市场与候选品初筛池.csv" and not args.dry_run:
        print(
            "执行失败：P07 后当前 03 已升级为中间候选样品池。"
            "请改用 scripts/build_candidate_pool.py 生成 runtime 03 / 60；"
            "本脚本只保留 market-only dry-run / cleaned report 用途。",
            file=sys.stderr,
        )
        return 1

    if args.dry_run:
        row_change_log = [
            {
                "action": "DRY_RUN_PREVIEW",
                "候选市场名称": row["候选市场名称"],
                "市场路径": row["市场路径"],
            }
            for row in mapped_rows
        ]
    else:
        try:
            headers, data_rows, row_change_log = merge_candidate_rows(
                repo_root=repo_root,
                candidate_csv_path=candidate_csv_path,
                mapped_rows=mapped_rows,
                append_mode=args.append,
            )
            write_csv_atomic(candidate_csv_path, headers, data_rows)
        except Exception as exc:
            print(f"执行失败：写入候选池 CSV 时出错：{exc}", file=sys.stderr)
            return 1

    report_json_path = output_dir / "market_to_candidate_pool_mapping_report.json"
    report_json = {
        "market_workbook": str(market_workbook),
        "workbook_selection_rule": WORKBOOK_SELECTION_RULE,
        "sheet_name": sheet_name,
        "header_row_index": header_row_index,
        "raw_headers": raw_headers,
        "row_reports": row_reports,
        "not_mapped_fields_union": sorted(not_mapped_fields_union),
        "row_change_log": row_change_log,
        "candidate_csv": str(candidate_csv_path),
        "cleaned_csv": str(cleaned_csv_path),
        "dry_run": args.dry_run,
        "mode": "append" if args.append else "replace_row",
    }
    report_json_path.write_text(json.dumps(report_json, ensure_ascii=False, indent=2), encoding="utf-8")

    report_markdown_path = output_dir / "market_to_candidate_pool_mapping_report.md"
    report_markdown_path.write_text(
        build_report_markdown(
            workbook_path=market_workbook,
            sheet_name=sheet_name,
            header_row_index=header_row_index,
            raw_headers=raw_headers,
            row_reports=row_reports,
            not_mapped_fields_union=sorted(not_mapped_fields_union),
            row_change_log=row_change_log,
            candidate_csv_path=candidate_csv_path,
            cleaned_csv_path=cleaned_csv_path,
            report_json_path=report_json_path,
            dry_run=args.dry_run,
            append_mode=args.append,
        ),
        encoding="utf-8",
    )

    print(f"市场表：{market_workbook}")
    print(f"识别工作表：{sheet_name}，标题行：{header_row_index}")
    print(f"清洗结果：{cleaned_csv_path}")
    print(f"映射报告：{report_markdown_path}")
    print(f"JSON报告：{report_json_path}")
    if args.dry_run:
        print("dry-run 已完成：未改写候选池 CSV。")
    else:
        print(f"候选池 CSV 已更新：{candidate_csv_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
