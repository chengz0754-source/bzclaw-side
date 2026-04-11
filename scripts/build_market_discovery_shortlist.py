from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from keyword_chain_common import ROOT, append_jsonl, ensure_within_repo, iso_now, write_json_atomic
from sellersprite_route_router import MARKET_DISCOVERY, resolve_route_decision


DEFAULT_LOG_DIR = ROOT / "logs" / "market_discovery_shortlist"
LATEST_FILE = "latest_run.json"
HISTORY_FILE = "market_discovery_shortlist_runs.jsonl"

CSV_READ_ENCODINGS = ("utf-8-sig", "utf-8", "gb18030", "gbk")
CURRENT_ROUTE_PATH = ROOT / "inputs" / "selection_run_current" / "01_选品任务路由与目的.csv"
CURRENT_MARKET_DISCOVERY_PATH = ROOT / "inputs" / "selection_run_current" / "01A_市场发现参数.csv"
DEFAULT_SELECTION_INPUT = ROOT / "inputs" / "selection_run_current" / "01__SELECTION_INPUT__TOY_10_TERMS_BATCH__20260411.csv"

SHORTLIST_CSV = "T01_市场发现短名单.csv"
SHORTLIST_MD = "T01_市场发现短名单.md"
SUMMARY_JSON = "T01_市场发现短名单_summary.json"

DEFAULT_WORKBOOK = Path(r"C:\Users\Administrator\Downloads\Market-research(200)SqueezeToys-US-Last-30-days (1).xlsx")


class MarketDiscoveryShortlistError(RuntimeError):
    def __init__(self, message: str, reason_code: str = "MARKET_DISCOVERY_SHORTLIST_ERROR") -> None:
        super().__init__(message)
        self.reason_code = reason_code


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build a T01 market discovery shortlist from either a real market workbook or a STEP3 gate projection.",
    )
    parser.add_argument("--context-row-index", type=int, default=1)
    parser.add_argument("--market-workbook", default=str(DEFAULT_WORKBOOK))
    parser.add_argument("--selection-input-csv", default="")
    parser.add_argument("--step3-cleaned-csv", default="")
    parser.add_argument("--step3-gate-csv", default="")
    parser.add_argument("--output-dir", required=True)
    parser.add_argument("--log-dir", default=str(DEFAULT_LOG_DIR))
    return parser.parse_args()


def repo_path(raw_path: str, label: str) -> Path:
    path = Path(raw_path).expanduser()
    if not path.is_absolute():
        path = ROOT / path
    return ensure_within_repo(path, label)


def read_text_with_fallback(path: Path) -> str:
    raw_bytes = path.read_bytes()
    decode_errors: list[str] = []
    for encoding in CSV_READ_ENCODINGS:
        try:
            return raw_bytes.decode(encoding)
        except UnicodeDecodeError as exc:
            decode_errors.append(f"{encoding}@{exc.start}:{exc.reason}")
    detail = " | ".join(decode_errors) or "unknown decode failure"
    raise MarketDiscoveryShortlistError(f"Failed to read CSV with supported encodings: {path} | {detail}")


def load_csv_rows(path: Path) -> list[list[str]]:
    return list(csv.reader(read_text_with_fallback(path).splitlines()))


def load_csv_dicts(path: Path) -> list[dict[str, str]]:
    return list(csv.DictReader(read_text_with_fallback(path).splitlines()))


def load_context_row(path: Path, row_index: int) -> dict[str, str]:
    rows = load_csv_rows(path)
    if len(rows) <= row_index:
        raise MarketDiscoveryShortlistError(
            f"--context-row-index {row_index} is out of range for {path}; available rows: {max(len(rows) - 1, 0)}",
            "CONTEXT_ROW_MISSING",
        )
    return {header: rows[row_index][idx] if idx < len(rows[row_index]) else "" for idx, header in enumerate(rows[0])}


def load_market_discovery_row(task_id: str) -> dict[str, str]:
    for row_map in load_csv_dicts(CURRENT_MARKET_DISCOVERY_PATH):
        if str(row_map.get("任务ID", "")).strip() == task_id:
            return row_map
    raise MarketDiscoveryShortlistError(
        f"No market-discovery row matched task_id={task_id!r} in {CURRENT_MARKET_DISCOVERY_PATH}",
        "MARKET_DISCOVERY_TASK_MISSING",
    )


def safe_float(value: Any) -> float:
    if value is None:
        return 0.0
    text = str(value).strip()
    if not text or text.upper() == "N/A":
        return 0.0
    if text.endswith("%"):
        text = text[:-1]
    try:
        return float(text)
    except (TypeError, ValueError):
        return 0.0


def parse_ratio(raw_value: Any) -> float:
    value = safe_float(raw_value)
    if value > 1:
        return value / 100.0
    return value


def parse_whitelist(raw_value: str) -> list[str]:
    return [item.strip() for item in str(raw_value or "").split("|") if item.strip()]


def parse_bool(value: str) -> bool:
    return str(value or "").strip().lower() in {"1", "true", "yes", "y", "是"}


def normalize_key(value: str) -> str:
    return " ".join(str(value or "").strip().casefold().split())


def format_percent(value: float) -> str:
    return f"{value * 100:.1f}%"


def format_number(value: float) -> str:
    return str(int(value)) if float(value).is_integer() else f"{value:.2f}"


def read_real_market_rows(workbook_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    parsed: list[dict[str, Any]] = []
    for row in worksheet.iter_rows(min_row=3, values_only=True):
        market_name = str(row[0] or "").strip()
        if not market_name:
            continue
        parsed.append(
            {
                "细分市场": market_name,
                "市场路径": str(row[2] or "").strip(),
                "月总销量": safe_float(row[4]),
                "月均销售额($)": safe_float(row[6]),
                "平均价格($)": safe_float(row[7]),
                "卖家集中度": parse_ratio(row[15]),
                "新品数量": safe_float(row[26]),
                "新品占比": parse_ratio(row[27]),
                "退货率": parse_ratio(row[33]),
                "搜索购买比": safe_float(row[35]),
            }
        )
    workbook.close()
    return parsed


def thresholds_from_market_row(row: dict[str, str]) -> dict[str, float]:
    return {
        "market_sales_floor": safe_float(row.get("市场容量下限")) or 300000.0,
        "new_product_ratio_floor": (safe_float(row.get("新品占比下限")) or 15.0) / 100.0,
        "seller_concentration_ceiling": (safe_float(row.get("卖家集中度上限")) or 80.0) / 100.0,
        "refund_rate_ceiling": (safe_float(row.get("退货率上限")) or 15.0) / 100.0,
        "price_floor": safe_float(row.get("价格带下限")) or 8.0,
        "price_ceiling": safe_float(row.get("价格带上限")) or 30.0,
    }


def decide_continue(row: dict[str, Any], thresholds: dict[str, float]) -> tuple[str, str]:
    path = str(row["市场路径"])
    if not path.startswith("Toys & Games:"):
        return "NO", "OUT_OF_SCOPE_CATEGORY_PATH"
    if row["月总销量"] < thresholds["market_sales_floor"]:
        return "NO", "MARKET_VOLUME_BELOW_FLOOR"
    if row["平均价格($)"] < thresholds["price_floor"] or row["平均价格($)"] > thresholds["price_ceiling"]:
        return "NO", "PRICE_OUT_OF_RANGE"
    if row["退货率"] > thresholds["refund_rate_ceiling"]:
        return "NO", "REFUND_RATE_TOO_HIGH"
    if row["卖家集中度"] > thresholds["seller_concentration_ceiling"]:
        return "HOLD", "SELLER_CONCENTRATION_TOO_HIGH"
    if row["新品占比"] >= thresholds["new_product_ratio_floor"]:
        return "YES", "VISIBLE_WHITELIST_MARKET_CONTINUE"
    if row["新品占比"] >= thresholds["new_product_ratio_floor"] - 0.03:
        return "HOLD", "NEW_PRODUCT_RATIO_NEAR_FLOOR"
    return "NO", "NEW_PRODUCT_RATIO_BELOW_FLOOR"


def load_selection_input_rows(path: Path) -> list[dict[str, str]]:
    rows = [row for row in load_csv_dicts(path) if parse_bool(row.get("enabled", "true"))]
    if not rows:
        raise MarketDiscoveryShortlistError(f"No enabled rows were found in selection input CSV: {path}", "SELECTION_INPUT_EMPTY")
    keys = [normalize_key(row.get("market_term", "")) for row in rows]
    if len(set(keys)) != len(keys):
        raise MarketDiscoveryShortlistError("Duplicate market_term rows detected in selection input CSV.", "SELECTION_INPUT_DUPLICATE_TERM")
    return sorted(rows, key=lambda row: (safe_float(row.get("priority") or 999), row.get("row_id", "")))


def load_rows_by_market_name(path: Path) -> dict[str, dict[str, str]]:
    rows: dict[str, dict[str, str]] = {}
    for row in load_csv_dicts(path):
        market_name = row.get("候选市场名称") or row.get("细分市场") or ""
        key = normalize_key(market_name)
        if key:
            rows[key] = row
    return rows


def gate_status_to_shortlist_status(step3_status: str) -> str:
    mapping = {"PASS": "YES", "HOLD": "HOLD", "FAIL": "NO"}
    return mapping.get(step3_status.strip().upper(), "NO")


def ranking_key(row: dict[str, str]) -> tuple[float, float, float, float]:
    shortlist_status = row["最终状态"]
    status_rank = {"YES": 0.0, "HOLD": 1.0, "NO": 2.0}.get(shortlist_status, 9.0)
    fail_count = safe_float(row.get("失败规则数", "999"))
    monthly_sales = safe_float(row.get("月总销量", "0"))
    priority = safe_float(row.get("priority", "999"))
    return (status_rank, fail_count, -monthly_sales, priority)


def recommend_terms(rows: list[dict[str, str]]) -> list[str]:
    yes_rows = [row for row in rows if row["最终状态"] == "YES"]
    if yes_rows:
        ordered = sorted(yes_rows, key=ranking_key)
        return [row["market_term"] for row in ordered[:2]]
    hold_rows = [row for row in rows if row["最终状态"] == "HOLD"]
    ordered = sorted(hold_rows, key=ranking_key)
    return [row["market_term"] for row in ordered[:2]]


def project_step3_gate_rows(
    selection_rows: list[dict[str, str]],
    cleaned_rows_by_market: dict[str, dict[str, str]],
    gate_rows_by_market: dict[str, dict[str, str]],
    task_id: str,
    route_decision: dict[str, Any],
) -> tuple[list[dict[str, str]], list[str]]:
    projected_rows: list[dict[str, str]] = []
    missing_terms: list[str] = []
    for selection_row in selection_rows:
        market_term = str(selection_row.get("market_term", "")).strip()
        market_key = normalize_key(market_term)
        gate_row = gate_rows_by_market.get(market_key)
        cleaned_row = cleaned_rows_by_market.get(market_key, {})
        if gate_row is None:
            missing_terms.append(market_term)
            projected_rows.append(
                {
                    "row_id": selection_row.get("row_id", ""),
                    "priority": selection_row.get("priority", ""),
                    "site": selection_row.get("site", ""),
                    "root_keyword": selection_row.get("root_keyword", ""),
                    "market_term": market_term,
                    "类目大类": selection_row.get("类目大类", ""),
                    "类目子类": selection_row.get("类目子类", ""),
                    "业务目的": selection_row.get("业务目的", ""),
                    "参数模板ID": selection_row.get("参数模板ID", ""),
                    "参数版本": selection_row.get("参数版本", ""),
                    "参数来源": selection_row.get("参数来源", ""),
                    "任务ID": task_id,
                    "purpose_type": str(route_decision.get("purpose_type", "")),
                    "候选市场名称": market_term,
                    "市场路径": "",
                    "平均价格": "",
                    "月总销量": "",
                    "新品占比_pct": "",
                    "商品集中度": "",
                    "品牌集中度": "",
                    "卖家集中度": "",
                    "命中规则数": "",
                    "失败规则数": "",
                    "STEP3整体状态": "MISSING",
                    "STEP3失败原因代码": "MARKET_TERM_NOT_FOUND_IN_STEP3",
                    "最终状态": "NO",
                    "shortlist_reason": "MARKET_TERM_NOT_FOUND_IN_STEP3",
                }
            )
            continue

        step3_status = str(gate_row.get("整体状态", "")).strip().upper()
        projected_rows.append(
            {
                "row_id": selection_row.get("row_id", ""),
                "priority": selection_row.get("priority", ""),
                "site": selection_row.get("site", ""),
                "root_keyword": selection_row.get("root_keyword", ""),
                "market_term": market_term,
                "类目大类": selection_row.get("类目大类", ""),
                "类目子类": selection_row.get("类目子类", ""),
                "业务目的": selection_row.get("业务目的", ""),
                "参数模板ID": selection_row.get("参数模板ID", ""),
                "参数版本": selection_row.get("参数版本", ""),
                "参数来源": selection_row.get("参数来源", ""),
                "任务ID": task_id,
                "purpose_type": str(route_decision.get("purpose_type", "")),
                "候选市场名称": gate_row.get("候选市场名称", market_term),
                "市场路径": cleaned_row.get("市场路径", cleaned_row.get("细分市场路径", "")),
                "平均价格": gate_row.get("平均价格", cleaned_row.get("平均价格", "")),
                "月总销量": gate_row.get("月总销量", cleaned_row.get("月总销量", "")),
                "新品占比_pct": gate_row.get("新品占比_pct", cleaned_row.get("新品占比_pct", "")),
                "商品集中度": gate_row.get("商品集中度", cleaned_row.get("商品集中度", "")),
                "品牌集中度": gate_row.get("品牌集中度", cleaned_row.get("品牌集中度", "")),
                "卖家集中度": gate_row.get("卖家集中度", cleaned_row.get("卖家集中度", "")),
                "命中规则数": gate_row.get("命中规则数", ""),
                "失败规则数": gate_row.get("失败规则数", ""),
                "STEP3整体状态": step3_status,
                "STEP3失败原因代码": gate_row.get("失败原因代码", "PASS"),
                "最终状态": gate_status_to_shortlist_status(step3_status),
                "shortlist_reason": "STEP3_GATE_PASS" if step3_status == "PASS" else gate_row.get("失败原因代码", ""),
            }
        )
    recommended = recommend_terms(projected_rows)
    recommended_set = {normalize_key(item) for item in recommended}
    for row in projected_rows:
        if normalize_key(row["market_term"]) in recommended_set:
            row["推荐动作"] = "RECOMMENDED_NEXT"
        elif row["最终状态"] == "YES":
            row["推荐动作"] = "SHORTLIST"
        elif row["最终状态"] == "HOLD":
            row["推荐动作"] = "WATCHLIST"
        else:
            row["推荐动作"] = "DROP"
    return projected_rows, missing_terms


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    ensure_within_repo(path, "shortlist_csv")
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fieldnames})


def markdown_summary(rows: list[dict[str, str]], summary: dict[str, Any]) -> str:
    lines = [
        "# T01 Market Discovery Shortlist",
        "",
        f"- Purpose: `{summary['purpose_type']}`",
        f"- Task ID: `{summary['task_id']}`",
        f"- Input term: `{summary['input_value']}`",
        f"- Mode: `{summary['mode']}`",
        f"- Selection input CSV: `{summary.get('selection_input_csv', '')}`",
        f"- Step3 gate CSV: `{summary.get('step3_gate_csv', '')}`",
        f"- Input rows: `{summary['input_row_count']}`",
        f"- Matched Step3 rows: `{summary['matched_step3_rows']}`",
        f"- Missing Step3 rows: `{summary['missing_step3_rows']}`",
        f"- YES rows: `{summary['continue_yes_rows']}`",
        f"- HOLD rows: `{summary['continue_hold_rows']}`",
        f"- NO rows: `{summary['continue_no_rows']}`",
        f"- Recommended terms: `{json.dumps(summary['recommended_terms'], ensure_ascii=False)}`",
        "",
        "| Market Term | Step3 Status | Final Status | Monthly Sales | Avg Price | Reason | Recommendation |",
        "| --- | --- | --- | --- | --- | --- | --- |",
    ]
    for row in rows:
        lines.append(
            f"| {row['market_term']} | {row['STEP3整体状态']} | {row['最终状态']} | {row['月总销量']} | {row['平均价格']} | {row['STEP3失败原因代码']} | {row['推荐动作']} |"
        )
    return "\n".join(lines) + "\n"


def persist_run(log_dir: Path, payload: dict[str, Any]) -> None:
    log_dir.mkdir(parents=True, exist_ok=True)
    write_json_atomic(log_dir / LATEST_FILE, payload)
    append_jsonl(log_dir / HISTORY_FILE, payload)


def run_legacy_workbook_mode(
    args: argparse.Namespace,
    output_dir: Path,
    log_dir: Path,
    route_row: dict[str, str],
    route_decision: dict[str, Any],
    task_id: str,
) -> dict[str, Any]:
    workbook_path = Path(args.market_workbook).expanduser()
    if not workbook_path.is_absolute():
        workbook_path = ROOT / workbook_path
    if not workbook_path.exists():
        raise MarketDiscoveryShortlistError(f"market workbook is missing: {workbook_path}", "MARKET_WORKBOOK_MISSING")

    market_row = load_market_discovery_row(task_id)
    whitelist = parse_whitelist(market_row.get("细分市场白名单", ""))
    thresholds = thresholds_from_market_row(market_row)
    visible_rows = read_real_market_rows(workbook_path)
    whitelist_rows = [row for row in visible_rows if str(row["细分市场"]) in whitelist]

    output_rows: list[dict[str, str]] = []
    for row in whitelist_rows:
        continue_flag, continue_reason = decide_continue(row, thresholds)
        output_rows.append(
            {
                "任务ID": task_id,
                "purpose_type": str(route_decision.get("purpose_type", "")),
                "market_term": str(row["细分市场"]),
                "候选市场名称": str(row["细分市场"]),
                "市场路径": str(row["市场路径"]),
                "月总销量": format_number(row["月总销量"]),
                "月均销售额($)": format_number(row["月均销售额($)"]),
                "平均价格": format_number(row["平均价格($)"]),
                "新品数量": format_number(row["新品数量"]),
                "新品占比_pct": format_percent(row["新品占比"]),
                "卖家集中度": format_percent(row["卖家集中度"]),
                "退货率": format_percent(row["退货率"]),
                "搜索购买比": format_number(row["搜索购买比"]),
                "STEP3整体状态": continue_flag,
                "STEP3失败原因代码": continue_reason,
                "最终状态": continue_flag,
                "shortlist_reason": continue_reason,
                "推荐动作": "SHORTLIST" if continue_flag == "YES" else "WATCHLIST" if continue_flag == "HOLD" else "DROP",
            }
        )

    summary = {
        "timestamp": iso_now(),
        "module": "build_market_discovery_shortlist",
        "status": "PASS",
        "reason_code": "PASS",
        "mode": "legacy_workbook",
        "task_id": task_id,
        "purpose_type": str(route_decision.get("purpose_type", "")),
        "input_value": str(route_row.get("input_value", "")),
        "market_workbook": str(workbook_path),
        "selection_input_csv": "",
        "step3_gate_csv": "",
        "input_row_count": len(whitelist),
        "matched_step3_rows": len(output_rows),
        "missing_step3_rows": 0,
        "continue_yes_rows": sum(1 for row in output_rows if row["最终状态"] == "YES"),
        "continue_hold_rows": sum(1 for row in output_rows if row["最终状态"] == "HOLD"),
        "continue_no_rows": sum(1 for row in output_rows if row["最终状态"] == "NO"),
        "recommended_terms": recommend_terms(output_rows),
        "output_csv": str(output_dir / SHORTLIST_CSV),
        "output_md": str(output_dir / SHORTLIST_MD),
        "output_summary": str(output_dir / SUMMARY_JSON),
    }
    write_outputs(output_dir, log_dir, output_rows, summary)
    return summary


def write_outputs(output_dir: Path, log_dir: Path, output_rows: list[dict[str, str]], summary: dict[str, Any]) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "row_id",
        "priority",
        "site",
        "root_keyword",
        "market_term",
        "类目大类",
        "类目子类",
        "业务目的",
        "参数模板ID",
        "参数版本",
        "参数来源",
        "任务ID",
        "purpose_type",
        "候选市场名称",
        "市场路径",
        "平均价格",
        "月总销量",
        "新品占比_pct",
        "商品集中度",
        "品牌集中度",
        "卖家集中度",
        "命中规则数",
        "失败规则数",
        "STEP3整体状态",
        "STEP3失败原因代码",
        "最终状态",
        "shortlist_reason",
        "推荐动作",
    ]
    write_csv(output_dir / SHORTLIST_CSV, fieldnames, output_rows)
    (output_dir / SHORTLIST_MD).write_text(markdown_summary(output_rows, summary), encoding="utf-8")
    write_json_atomic(output_dir / SUMMARY_JSON, summary)
    persist_run(log_dir, summary)


def run_gate_projection_mode(
    args: argparse.Namespace,
    output_dir: Path,
    log_dir: Path,
    route_row: dict[str, str],
    route_decision: dict[str, Any],
    task_id: str,
) -> dict[str, Any]:
    selection_input_csv = repo_path(args.selection_input_csv or str(DEFAULT_SELECTION_INPUT), "selection_input_csv")
    step3_gate_csv = repo_path(args.step3_gate_csv, "step3_gate_csv")
    cleaned_csv = repo_path(args.step3_cleaned_csv, "step3_cleaned_csv") if args.step3_cleaned_csv else None
    selection_rows = load_selection_input_rows(selection_input_csv)
    gate_rows_by_market = load_rows_by_market_name(step3_gate_csv)
    cleaned_rows_by_market = load_rows_by_market_name(cleaned_csv) if cleaned_csv else {}

    projected_rows, missing_terms = project_step3_gate_rows(
        selection_rows=selection_rows,
        cleaned_rows_by_market=cleaned_rows_by_market,
        gate_rows_by_market=gate_rows_by_market,
        task_id=task_id,
        route_decision=route_decision,
    )
    recommended = recommend_terms(projected_rows)

    summary = {
        "timestamp": iso_now(),
        "module": "build_market_discovery_shortlist",
        "status": "PASS",
        "reason_code": "PASS",
        "mode": "step3_gate_projection",
        "task_id": task_id,
        "purpose_type": str(route_decision.get("purpose_type", "")),
        "input_value": str(route_row.get("input_value", "")),
        "market_workbook": "",
        "selection_input_csv": str(selection_input_csv),
        "step3_gate_csv": str(step3_gate_csv),
        "step3_cleaned_csv": str(cleaned_csv) if cleaned_csv else "",
        "input_row_count": len(selection_rows),
        "matched_step3_rows": len(projected_rows) - len(missing_terms),
        "missing_step3_rows": len(missing_terms),
        "missing_terms": missing_terms,
        "continue_yes_rows": sum(1 for row in projected_rows if row["最终状态"] == "YES"),
        "continue_hold_rows": sum(1 for row in projected_rows if row["最终状态"] == "HOLD"),
        "continue_no_rows": sum(1 for row in projected_rows if row["最终状态"] == "NO"),
        "recommended_terms": recommended,
        "status_distribution": {
            "YES": sum(1 for row in projected_rows if row["最终状态"] == "YES"),
            "HOLD": sum(1 for row in projected_rows if row["最终状态"] == "HOLD"),
            "NO": sum(1 for row in projected_rows if row["最终状态"] == "NO"),
        },
        "output_csv": str(output_dir / SHORTLIST_CSV),
        "output_md": str(output_dir / SHORTLIST_MD),
        "output_summary": str(output_dir / SUMMARY_JSON),
    }
    write_outputs(output_dir, log_dir, projected_rows, summary)
    return summary


def main() -> int:
    args = parse_args()
    output_dir = repo_path(args.output_dir, "output_dir")
    log_dir = repo_path(args.log_dir, "log_dir")

    route_row = load_context_row(CURRENT_ROUTE_PATH, args.context_row_index)
    task_id = str(route_row.get("任务ID", "")).strip()
    route_decision = resolve_route_decision(context_row_index=args.context_row_index)
    if str(route_decision.get("purpose_type", "")) != MARKET_DISCOVERY:
        raise MarketDiscoveryShortlistError(
            f"Context row {args.context_row_index} is not MARKET_DISCOVERY; got {route_decision.get('purpose_type')!r}",
            "PURPOSE_NOT_MARKET_DISCOVERY",
        )

    if args.step3_gate_csv:
        summary = run_gate_projection_mode(args, output_dir, log_dir, route_row, route_decision, task_id)
    else:
        summary = run_legacy_workbook_mode(args, output_dir, log_dir, route_row, route_decision, task_id)
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
