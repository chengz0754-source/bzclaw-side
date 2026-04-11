from __future__ import annotations

import argparse
import json
import re
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
from playwright.sync_api import sync_playwright

from benchmark_chain_common import (
    BENCHMARK_RAW_ARTIFACT,
    PROFILE_DIR,
    BenchmarkChainError,
    clean_number,
    ensure_within_repo,
    iso_now,
    log_dir_from_namespace,
    output_dir_from_namespace,
    persist_run_summary,
    preferred_sellersprite_profile_dir,
    resolve_context_from_namespace,
    resolve_seed_from_upstream,
)
from keyword_chain_common import REPLAY_PROFILE_DIR, STORAGE_STATE_PATH, compact_text, write_json_atomic
from sellersprite_auth_registry import register_auth_incident, replay_meta_from_incident
from sellersprite_auth_replay import launch_runtime_seeded_persistent_context, perform_registered_login_replay, summary_requests_auth_replay
from parse_benchmark_export_workbook import build_raw_artifact, parse_workbook_rows
from sellersprite_overlay_guard import capture_screenshot, find_first_visible, guard_page, page_identity


try:
    sys.stdout.reconfigure(encoding="utf-8", errors="backslashreplace")
    sys.stderr.reconfigure(encoding="utf-8", errors="backslashreplace")
except Exception:
    pass


ROOT = Path(__file__).resolve().parents[1]
BENCHMARK_URL = "https://www.sellersprite.com/v3/competitor-lookup"
EXPORT_LOG_URL = "https://www.sellersprite.com/v2/export-log"
SCREENSHOT_DIR = ROOT / "playwright" / "screenshots" / "benchmark_chain"
DOWNLOADS_ROOT = ROOT / "runs" / "manual" / "20_benchmark_exports"

PAGE_OPEN_WAIT_MS = 2200
PRE_CLICK_WAIT_MS = 700
AFTER_QUERY_WAIT_MS = 3000
AFTER_EXPORT_WAIT_MS = 2500
CLICK_RETRY_WAIT_MS = 1200
MIN_POLL_INTERVAL_SECONDS = 8

RESULT_PAGE_BLOCKED_BY_OVERLAY = "RESULT_PAGE_BLOCKED_BY_OVERLAY"
EXPORT_DIALOG_NOT_VISIBLE = "EXPORT_DIALOG_NOT_VISIBLE"
EXPORT_CONFIRM_BUTTON_NOT_VISIBLE = "EXPORT_CONFIRM_BUTTON_NOT_VISIBLE"
EXPORT_LOG_TASK_NOT_FOUND = "EXPORT_LOG_TASK_NOT_FOUND"
EXPORT_LOG_STATUS_TIMEOUT = "EXPORT_LOG_STATUS_TIMEOUT"
EXPORT_DOWNLOAD_BUTTON_NOT_VISIBLE = "EXPORT_DOWNLOAD_BUTTON_NOT_VISIBLE"
EXPORT_FILE_NOT_DOWNLOADED = "EXPORT_FILE_NOT_DOWNLOADED"
UNEXPECTED_MODAL_BLOCKING_ACTION = "UNEXPECTED_MODAL_BLOCKING_ACTION"
SELLERSPRITE_AUTH_REQUIRED = "SELLERSPRITE_AUTH_REQUIRED"

BENCHMARK_QUERY_INPUT_NOT_VISIBLE = "BENCHMARK_QUERY_INPUT_NOT_VISIBLE"
BENCHMARK_QUERY_BUTTON_NOT_VISIBLE = "BENCHMARK_QUERY_BUTTON_NOT_VISIBLE"
BENCHMARK_RESULT_TABLE_NOT_VISIBLE = "BENCHMARK_RESULT_TABLE_NOT_VISIBLE"
BENCHMARK_RESULT_ROW_NOT_FOUND = "BENCHMARK_RESULT_ROW_NOT_FOUND"
BENCHMARK_RESULT_CHECKBOX_NOT_VISIBLE = "BENCHMARK_RESULT_CHECKBOX_NOT_VISIBLE"
EXPORT_LOG_STATUS_FAILED = "EXPORT_LOG_STATUS_FAILED"
EXPORT_WORKBOOK_PARSE_FAILED = "EXPORT_WORKBOOK_PARSE_FAILED"

SITE_LABELS = {
    "US": "美国站",
    "JP": "日本站",
    "UK": "英国站",
    "DE": "德国站",
    "FR": "法国站",
    "IT": "意大利站",
    "ES": "西班牙站",
    "CA": "加拿大站",
    "IN": "印度站",
    "MX": "墨西哥站",
}
LAST_30_DAYS_LABEL = "最近30天"

EXPORT_DIALOG_CONFIRM_SELECTORS = (
    "button:has-text('前往查看')",
    "a:has-text('前往查看')",
    "button:has-text('去查看')",
    "a:has-text('去查看')",
    "button:has-text('查看')",
    "a:has-text('查看')",
    "button:has-text('我的导出')",
    "a:has-text('我的导出')",
)

EXPORT_DIALOG_SELECTORS = (
    ".el-message-box",
    ".modal.show",
    ".el-dialog",
    "div[role='dialog']",
)

STATUS_DONE_MARKERS = ("已完成", "完成", "success", "done")
STATUS_PROGRESS_MARKERS = ("导出中", "处理中", "生成中", "等待", "排队", "pending", "processing")
STATUS_FAIL_MARKERS = ("失败", "过期", "取消", "异常", "error", "failed")


class BenchmarkExportFlowError(RuntimeError):
    def __init__(self, message: str, reason_code: str) -> None:
        super().__init__(message)
        self.reason_code = reason_code


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export SellerSprite benchmark results through the page export-log workflow, then parse the downloaded workbook into STEP4 raw artifact.",
    )
    parser.add_argument("--context-row-index", type=int, default=1)
    parser.add_argument("--run-name", default=None)
    parser.add_argument("--direction-id", default=None)
    parser.add_argument("--keyword", default=None)
    parser.add_argument("--category-hint", default=None)
    parser.add_argument("--site", default=None)
    parser.add_argument("--days", type=int, default=None)
    parser.add_argument("--sample-top-n", type=int, default=None)
    parser.add_argument("--max-candidate-samples", type=int, default=None)
    parser.add_argument("--seed-keyword", default=None, help="Debug-only explicit seed keyword override.")
    parser.add_argument("--seed-market-name", default=None, help="Optional market name when --seed-keyword is supplied.")
    parser.add_argument("--product-gate-csv", default=None)
    parser.add_argument("--product-seed-csv", default=None)
    parser.add_argument("--step3-gate-csv", default=None)
    parser.add_argument("--step3-cleaned-csv", default=None)
    parser.add_argument("--output-dir", default=None)
    parser.add_argument("--log-dir", default=None)
    parser.add_argument("--download-dir", default=None)
    parser.add_argument("--result-row-index", type=int, default=1)
    parser.add_argument("--max-wait-seconds", type=int, default=300)
    parser.add_argument("--poll-interval-seconds", type=int, default=10)
    parser.add_argument("--download-timeout-seconds", type=int, default=90)
    parser.add_argument("--headless", action="store_true", help="Use headless mode. Current verified live path prefers headed mode.")
    parser.add_argument("--dry-run", dest="dry_run", action="store_true")
    parser.add_argument("--execution-mode", choices=("auto", "persistent", "storage_state"), default="auto")
    return parser.parse_args()


def slugify(value: str) -> str:
    slug = re.sub(r"[^A-Za-z0-9]+", "-", str(value or "").strip()).strip("-")
    return slug or "benchmark-export"


def wait_for_page_open(page) -> None:
    page.wait_for_timeout(PAGE_OPEN_WAIT_MS)


def wait_before_click(page) -> None:
    page.wait_for_timeout(PRE_CLICK_WAIT_MS)


def wait_after_query(page) -> None:
    page.wait_for_timeout(AFTER_QUERY_WAIT_MS)


def wait_after_export(page) -> None:
    page.wait_for_timeout(AFTER_EXPORT_WAIT_MS)


def ensure_min_poll_interval(seconds: int) -> int:
    return max(MIN_POLL_INTERVAL_SECONDS, int(seconds))


def safe_download_dir(args: argparse.Namespace) -> Path:
    if args.download_dir:
        download_dir = Path(args.download_dir).expanduser()
        if not download_dir.is_absolute():
            download_dir = ROOT / download_dir
        return ensure_within_repo(download_dir, "download_dir")
    return ensure_within_repo(DOWNLOADS_ROOT / datetime.now().strftime("%Y%m%d_%H%M%S"), "download_dir")


def login_required(page) -> bool:
    if "/w/user/login" in page.url:
        return True
    snapshot = page_identity(page)
    title = str(snapshot.get("title", ""))
    if "登录" in title:
        return True
    return bool(snapshot.get("guest_markers"))


def selected_site_value(page) -> str:
    return page.locator(".market-select input").first.input_value().strip()


def selected_month_value(page) -> str:
    return page.locator(".fliter .el-select").nth(1).locator("input").first.input_value().strip()


def choose_dropdown_value(page, select_locator, target_text: str) -> None:
    current_value = select_locator.locator("input").first.input_value().strip()
    if current_value == target_text:
        return
    wait_before_click(page)
    select_locator.locator("input").first.click(timeout=10000)
    option = page.locator(".el-select-dropdown__item", has_text=target_text).last
    option.click(timeout=10000)
    page.wait_for_timeout(800)


def configure_filters(page, site_code: str, days: int) -> dict[str, str]:
    if site_code not in SITE_LABELS:
        raise BenchmarkChainError(f"Unsupported benchmark site code: {site_code}", "BENCHMARK_SITE_UNSUPPORTED")
    if days != 30:
        raise BenchmarkChainError(
            "The current benchmark export chain is only verified against 最近30天. Use 时间范围_天=30 for now.",
            "BENCHMARK_DAYS_UNSUPPORTED",
        )

    choose_dropdown_value(page, page.locator(".market-select").first, SITE_LABELS[site_code])
    choose_dropdown_value(page, page.locator(".fliter .el-select").nth(1), LAST_30_DAYS_LABEL)
    return {
        "site_label": selected_site_value(page),
        "month_label": selected_month_value(page),
    }


def benchmark_query_input(page):
    return page.locator(".filter-item.input-wrap input[placeholder*='flashlight']").first


def benchmark_query_button(page):
    return page.locator(".filter-item.input-wrap button").first


def record_step(summary: dict[str, Any], step_name: str, status: str, page=None, reason_code: str = "", screenshot_path: str = "", extra: dict[str, Any] | None = None) -> None:
    payload: dict[str, Any] = {
        "timestamp": iso_now(),
        "step_name": step_name,
        "status": status,
        "reason_code": reason_code,
    }
    if page is not None:
        snapshot = page_identity(page)
        payload["page_url"] = snapshot.get("url", "")
        payload["page_title"] = snapshot.get("title", "")
    if screenshot_path:
        payload["screenshot_path"] = screenshot_path
    if extra:
        payload.update(extra)
    summary.setdefault("steps", []).append(payload)
    summary["current_step"] = step_name


def fail_closed_on_auth(page, summary: dict[str, Any], step_name: str, message: str, redirect_from_url: str, run_context: dict[str, Any]) -> None:
    incident = register_auth_incident(
        module_name="benchmark_export",
        step_name=step_name,
        source_script=__file__,
        reason_code=SELLERSPRITE_AUTH_REQUIRED,
        current_url=page.url if page is not None else "",
        redirect_from_url=redirect_from_url,
        page=page,
        run_context=run_context,
    )
    summary.update(replay_meta_from_incident(incident))
    record_step(
        summary,
        step_name,
        "FAIL",
        page=page,
        reason_code=SELLERSPRITE_AUTH_REQUIRED,
        screenshot_path=str(incident.get("screenshot_path", "")).strip(),
        extra={
            "auth_surface_family": incident.get("surface_family", ""),
            "auth_replay_available": incident.get("has_login_replay", False),
        },
    )
    raise BenchmarkExportFlowError(message, SELLERSPRITE_AUTH_REQUIRED)


def launch_context(
    playwright,
    args: argparse.Namespace,
    *,
    runtime_replay_surface_family: str = "SELLERSPRITE_COMPETITOR_LOOKUP_AUTH",
) -> tuple[Any, Any, str, str]:
    warning = ""
    browser = None
    persistent_profile_dir = preferred_sellersprite_profile_dir()
    if args.execution_mode in {"auto", "persistent"} and persistent_profile_dir is not None:
        launch_errors: list[str] = []
        for attempt_index in range(1, 3):
            try:
                if persistent_profile_dir == REPLAY_PROFILE_DIR:
                    context_browser, runtime_info = launch_runtime_seeded_persistent_context(
                        playwright,
                        surface_family=runtime_replay_surface_family,
                        headless=bool(args.headless),
                        viewport={"width": 1600, "height": 1400},
                        accept_downloads=True,
                        channel="msedge",
                    )
                    warning = ((warning + " | ") if warning else "") + (
                        f"using_runtime_replay_surface={runtime_replay_surface_family}; "
                        f"runtime_profile_dir={runtime_info.get('runtime_profile_dir', '')}"
                    )
                else:
                    context_browser = playwright.chromium.launch_persistent_context(
                        str(persistent_profile_dir),
                        channel="msedge",
                        headless=bool(args.headless),
                        viewport={"width": 1600, "height": 1400},
                        accept_downloads=True,
                    )
                return context_browser, browser, "persistent_profile", warning
            except Exception as exc:
                launch_errors.append(f"attempt_{attempt_index}:{exc}")
                if attempt_index == 1:
                    time.sleep(1.5)
                    continue
                if args.execution_mode == "persistent":
                    raise
                warning = "; ".join(launch_errors)

    browser = playwright.chromium.launch(channel="msedge", headless=bool(args.headless))
    if STORAGE_STATE_PATH.exists():
        context_browser = browser.new_context(
            storage_state=str(STORAGE_STATE_PATH),
            viewport={"width": 1600, "height": 1400},
            accept_downloads=True,
        )
        return context_browser, browser, "storage_state", warning
    context_browser = browser.new_context(viewport={"width": 1600, "height": 1400}, accept_downloads=True)
    return context_browser, browser, "guest_context", warning


def require_visible(page, locator, step_name: str, summary: dict[str, Any], not_visible_reason: str, preserve_texts: tuple[str, ...] = ()) -> Any:
    guard = guard_page(page, f"{step_name}-guard", preserve_texts=preserve_texts)
    try:
        locator.first.wait_for(state="visible", timeout=10000)
        return locator.first
    except Exception as exc:
        screenshot_path = capture_screenshot(page, f"{step_name}-not-visible", SCREENSHOT_DIR)
        record_step(
            summary,
            step_name,
            "FAIL",
            page=page,
            reason_code=not_visible_reason,
            screenshot_path=screenshot_path,
            extra={"guard": guard, "error": str(exc)},
        )
        raise BenchmarkExportFlowError(f"{step_name} target is not visible.", not_visible_reason) from exc


def ensure_query_surface(page, summary: dict[str, Any]) -> None:
    require_visible(
        page,
        benchmark_query_input(page),
        "stage_a_open_query_surface",
        summary,
        BENCHMARK_QUERY_INPUT_NOT_VISIBLE,
        preserve_texts=("立即查询",),
    )
    require_visible(
        page,
        benchmark_query_button(page),
        "stage_a_open_query_surface_button",
        summary,
        BENCHMARK_QUERY_BUTTON_NOT_VISIBLE,
        preserve_texts=("立即查询",),
    )


def trigger_query(page, keyword: str, summary: dict[str, Any]) -> None:
    query_input = require_visible(
        page,
        benchmark_query_input(page),
        "stage_a_fill_query_keyword",
        summary,
        BENCHMARK_QUERY_INPUT_NOT_VISIBLE,
        preserve_texts=("立即查询",),
    )
    wait_before_click(page)
    query_input.click(timeout=10000)
    query_input.fill(keyword)
    query_button = require_visible(
        page,
        benchmark_query_button(page),
        "stage_a_submit_query",
        summary,
        BENCHMARK_QUERY_BUTTON_NOT_VISIBLE,
        preserve_texts=("立即查询",),
    )
    wait_before_click(page)
    query_button.click(timeout=10000)
    wait_after_query(page)


def result_row_checkbox(page, row_index: int):
    return page.locator(".el-table__body .el-checkbox.table-check").nth(max(0, row_index - 1))


def result_bulk_checkbox(page):
    return page.locator("div.left > label.el-checkbox").first


def result_rows(page):
    return page.locator(".el-table__body tr")


def locate_export_button(page):
    button = page.locator("button.my-download")
    if button.count():
        return button.first
    return None


def checkbox_selected(locator) -> bool:
    try:
        return "is-checked" in str(locator.get_attribute("class") or "")
    except Exception:
        return False


def select_result_row(page, row_index: int, summary: dict[str, Any]) -> None:
    selection_scope = "ALL_VISIBLE_RESULTS"
    try:
        checkbox = require_visible(
            page,
            result_bulk_checkbox(page),
            "stage_a_select_result_row",
            summary,
            BENCHMARK_RESULT_CHECKBOX_NOT_VISIBLE,
            preserve_texts=("导出", "导出明细"),
        )
    except BenchmarkExportFlowError:
        selection_scope = f"ROW_{row_index}"
        checkbox = require_visible(
            page,
            result_row_checkbox(page, row_index),
            "stage_a_select_result_row",
            summary,
            BENCHMARK_RESULT_CHECKBOX_NOT_VISIBLE,
            preserve_texts=("导出", "导出明细"),
        )

    for attempt in range(1, 3):
        guard = guard_page(page, f"stage_a_select_result_row-attempt-{attempt}", preserve_texts=("导出", "导出明细"))
        wait_before_click(page)
        checkbox.click(timeout=10000, force=True)
        page.wait_for_timeout(1000)
        if checkbox_selected(checkbox):
            record_step(
                summary,
                "stage_a_select_result_row",
                "PASS",
                page=page,
                extra={"attempt": attempt, "guard": guard, "selection_scope": selection_scope},
            )
            return
        if attempt < 2:
            page.wait_for_timeout(CLICK_RETRY_WAIT_MS)

    screenshot_path = capture_screenshot(page, "stage_a_select_result_row-failure", SCREENSHOT_DIR)
    record_step(
        summary,
        "stage_a_select_result_row",
        "FAIL",
        page=page,
        reason_code=RESULT_PAGE_BLOCKED_BY_OVERLAY,
        screenshot_path=screenshot_path,
        extra={"attempts": 2},
    )
    raise BenchmarkExportFlowError("Benchmark result-row checkbox could not be selected.", RESULT_PAGE_BLOCKED_BY_OVERLAY)


def handle_optional_export_dialog(page) -> dict[str, Any]:
    dialog = find_first_visible(page, EXPORT_DIALOG_SELECTORS, timeout_ms=1000)
    confirm = find_first_visible(page, EXPORT_DIALOG_CONFIRM_SELECTORS, timeout_ms=1000)
    if dialog is None and confirm is None:
        return {"dialog_visible": False, "confirm_clicked": False}

    if confirm is None:
        raise BenchmarkExportFlowError("Benchmark export dialog appeared, but no stable confirm button was found.", EXPORT_CONFIRM_BUTTON_NOT_VISIBLE)

    try:
        with page.expect_popup(timeout=5000) as popup_info:
            wait_before_click(page)
            confirm.click(timeout=10000)
        popup = popup_info.value
        wait_after_export(popup)
        return {
            "dialog_visible": True,
            "confirm_clicked": True,
            "via": "popup",
            "popup_url": popup.url,
        }
    except PlaywrightTimeoutError:
        wait_before_click(page)
        confirm.click(timeout=10000)
        wait_after_export(page)
        return {
            "dialog_visible": True,
            "confirm_clicked": True,
            "via": "same_tab",
            "popup_url": "",
        }


def trigger_export(page, summary: dict[str, Any]) -> dict[str, Any]:
    for attempt in range(1, 3):
        guard = guard_page(page, f"stage_a_trigger_export-attempt-{attempt}", preserve_texts=("导出", "导出明细"))
        button = locate_export_button(page)
        if button is None:
            screenshot_path = capture_screenshot(page, "stage_a_trigger_export-missing", SCREENSHOT_DIR)
            record_step(
                summary,
                "stage_a_trigger_export",
                "FAIL",
                page=page,
                reason_code=EXPORT_DIALOG_NOT_VISIBLE,
                screenshot_path=screenshot_path,
                extra={"guard": guard},
            )
            raise BenchmarkExportFlowError("Benchmark export button is not visible.", EXPORT_DIALOG_NOT_VISIBLE)

        wait_before_click(page)
        try:
            button.click(timeout=10000)
        except Exception as exc:
            if attempt < 2:
                page.wait_for_timeout(CLICK_RETRY_WAIT_MS)
                continue
            screenshot_path = capture_screenshot(page, "stage_a_trigger_export-failure", SCREENSHOT_DIR)
            record_step(
                summary,
                "stage_a_trigger_export",
                "FAIL",
                page=page,
                reason_code=RESULT_PAGE_BLOCKED_BY_OVERLAY,
                screenshot_path=screenshot_path,
                extra={"guard": guard, "attempt": attempt, "error": str(exc)},
            )
            raise BenchmarkExportFlowError("Benchmark export button remained blocked.", RESULT_PAGE_BLOCKED_BY_OVERLAY) from exc

        wait_after_export(page)
        follow_action = handle_optional_export_dialog(page)
        record_step(
            summary,
            "stage_a_trigger_export",
            "PASS",
            page=page,
            extra={"attempt": attempt, "guard": guard, "follow_action": follow_action},
        )
        return follow_action

    raise BenchmarkExportFlowError("Benchmark export button could not be triggered.", RESULT_PAGE_BLOCKED_BY_OVERLAY)


def parse_export_log_time(value: str) -> float | None:
    text = compact_text(value)
    if not text or text == "-":
        return None
    try:
        return datetime.strptime(text, "%Y-%m-%d %H:%M").timestamp()
    except ValueError:
        return None


def collect_export_tasks(page) -> list[dict[str, Any]]:
    rows = page.locator("tbody tr")
    tasks: list[dict[str, Any]] = []
    for index in range(rows.count()):
        row = rows.nth(index)
        cells = row.locator("td")
        if cells.count() < 6:
            continue
        task_name = compact_text(cells.nth(1).inner_text(timeout=3000)).replace(" 最新", "").strip()
        source_name = compact_text(cells.nth(2).inner_text(timeout=3000)).strip()
        time_lines = cells.nth(3).inner_text(timeout=3000).splitlines()
        generated_at = compact_text(time_lines[0] if len(time_lines) >= 1 else "")
        download_at = compact_text(time_lines[1] if len(time_lines) >= 2 else "")
        status_value = compact_text(cells.nth(4).inner_text(timeout=3000)).strip()
        download_href = ""
        download_button = row.locator("a.download-excel, button.download-excel, .download-excel")
        if download_button.count():
            download_href = str(download_button.first.get_attribute("href") or "").strip()
        tasks.append(
            {
                "row_index": index,
                "task_name": task_name,
                "source_name": source_name,
                "generated_at": generated_at,
                "generated_ts": parse_export_log_time(generated_at),
                "download_at": download_at,
                "status_value": status_value,
                "download_href": download_href,
            }
        )
    return tasks


def expected_task_prefix(site: str, days: int) -> str:
    return f"Competitor-{site}-Last-{days}-days"


def find_best_export_task(tasks: list[dict[str, Any]], baseline_task_names: set[str], expected_prefix: str, export_triggered_ts: float) -> dict[str, Any] | None:
    ranked: list[tuple[int, float, dict[str, Any]]] = []
    for task in tasks:
        task_name = str(task.get("task_name", "")).strip()
        source_name = str(task.get("source_name", "")).strip()
        if not task_name:
            continue
        if not task_name.startswith("Competitor-") and "查竞品" not in source_name:
            continue

        score = 0
        if task_name.startswith(expected_prefix):
            score += 500
        if "查竞品" in source_name:
            score += 200
        if task_name not in baseline_task_names:
            score += 1000
        generated_ts = task.get("generated_ts")
        if isinstance(generated_ts, (int, float)):
            delta = abs(float(generated_ts) - export_triggered_ts)
            if delta <= 300:
                score += 200
            elif delta <= 900:
                score += 100
        if score <= 0:
            continue
        ranked.append((score, float(generated_ts or 0.0), task))

    if not ranked:
        return None
    ranked.sort(key=lambda item: (item[0], item[1]), reverse=True)
    return ranked[0][2]


def export_status_kind(status_value: str) -> str:
    lowered = compact_text(status_value).casefold()
    if any(marker.casefold() in lowered for marker in STATUS_DONE_MARKERS):
        return "DONE"
    if any(marker.casefold() in lowered for marker in STATUS_FAIL_MARKERS):
        return "FAIL"
    if any(marker.casefold() in lowered for marker in STATUS_PROGRESS_MARKERS):
        return "WAIT"
    return "UNKNOWN"


def safe_download_filename(task_name: str, suggested_filename: str) -> str:
    suffix = Path(suggested_filename or "").suffix.lower() or ".xlsx"
    stem = Path(suggested_filename or "").stem or task_name
    safe_stem = re.sub(r"[^A-Za-z0-9._-]+", "-", stem).strip("-") or slugify(task_name)
    return f"{safe_stem}{suffix}"


def validate_downloaded_workbook(path: Path, expected_prefix: str) -> dict[str, Any]:
    if not path.exists():
        raise BenchmarkExportFlowError(f"Benchmark export workbook was not saved: {path}", EXPORT_FILE_NOT_DOWNLOADED)
    if path.stat().st_size <= 0:
        raise BenchmarkExportFlowError(f"Benchmark export workbook is empty: {path}", EXPORT_FILE_NOT_DOWNLOADED)
    suffix = path.suffix.lower()
    if suffix not in {".xlsx", ".xls"}:
        raise BenchmarkExportFlowError(f"Benchmark export workbook suffix is unexpected: {path.name}", EXPORT_FILE_NOT_DOWNLOADED)
    if expected_prefix.casefold() not in path.stem.casefold():
        raise BenchmarkExportFlowError(
            f"Benchmark export workbook filename does not match expected prefix `{expected_prefix}`: {path.name}",
            EXPORT_FILE_NOT_DOWNLOADED,
        )
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet_names = workbook.sheetnames
        workbook.close()
    except Exception as exc:
        raise BenchmarkExportFlowError(
            f"Benchmark export workbook could not be opened by openpyxl: {path.name}",
            EXPORT_FILE_NOT_DOWNLOADED,
        ) from exc
    return {
        "path": str(path),
        "name": path.name,
        "size_bytes": path.stat().st_size,
        "sheet_names": sheet_names,
    }


def download_task_workbook(page, task: dict[str, Any], download_dir: Path, timeout_ms: int, summary: dict[str, Any]) -> tuple[Path, dict[str, Any]]:
    task_name = str(task.get("task_name", "")).strip()
    matching_row = page.locator("tbody tr").filter(has_text=task_name).first
    try:
        matching_row.wait_for(state="visible", timeout=10000)
    except Exception as exc:
        screenshot_path = capture_screenshot(page, "stage_b_download_export-row-missing", SCREENSHOT_DIR)
        record_step(
            summary,
            "stage_b_download_export",
            "FAIL",
            page=page,
            reason_code=EXPORT_DOWNLOAD_BUTTON_NOT_VISIBLE,
            screenshot_path=screenshot_path,
            extra={"task_name": task_name, "error": str(exc)},
        )
        raise BenchmarkExportFlowError("Matched benchmark export task row is no longer visible.", EXPORT_DOWNLOAD_BUTTON_NOT_VISIBLE) from exc

    button = matching_row.locator("a.download-excel, button.download-excel, .download-excel").first
    try:
        button.wait_for(state="visible", timeout=10000)
    except Exception as exc:
        screenshot_path = capture_screenshot(page, "stage_b_download_export-button-missing", SCREENSHOT_DIR)
        record_step(
            summary,
            "stage_b_download_export",
            "FAIL",
            page=page,
            reason_code=EXPORT_DOWNLOAD_BUTTON_NOT_VISIBLE,
            screenshot_path=screenshot_path,
            extra={"task_name": task_name, "error": str(exc)},
        )
        raise BenchmarkExportFlowError("Benchmark export download button is not visible in the matched row.", EXPORT_DOWNLOAD_BUTTON_NOT_VISIBLE) from exc

    wait_before_click(page)
    with page.expect_download(timeout=timeout_ms) as download_info:
        button.click(timeout=10000)
    download = download_info.value
    download_dir.mkdir(parents=True, exist_ok=True)
    target_path = ensure_within_repo(download_dir / safe_download_filename(task_name, download.suggested_filename), "download_target")
    download.save_as(str(target_path))
    file_info = validate_downloaded_workbook(target_path, "Competitor-")
    record_step(
        summary,
        "stage_b_download_export",
        "PASS",
        page=page,
        extra={"task_name": task_name, "download": file_info},
    )
    return target_path, file_info


def resolve_seed(args: argparse.Namespace, context) -> dict[str, Any]:
    if args.seed_keyword:
        seed_keyword = str(args.seed_keyword).strip()
        if not seed_keyword:
            raise BenchmarkChainError("--seed-keyword cannot be blank.", "SEED_KEYWORD_BLANK")
        return {
            "source_step": "MANUAL_OVERRIDE",
            "source_gate_path": "",
            "source_cleaned_path": "",
            "seed_keyword": seed_keyword,
            "candidate_market_name": str(args.seed_market_name or seed_keyword).strip(),
            "market_path": "",
            "upstream_batch_id": "",
            "upstream_status": "MANUAL_OVERRIDE",
        }

    resolved_seed = resolve_seed_from_upstream(
        context,
        args.product_gate_csv,
        args.product_seed_csv,
        args.step3_gate_csv,
        args.step3_cleaned_csv,
    )
    return {
        "source_step": resolved_seed.source_step,
        "source_gate_path": resolved_seed.source_gate_path,
        "source_cleaned_path": resolved_seed.source_cleaned_path,
        "seed_keyword": resolved_seed.seed_keyword,
        "candidate_market_name": resolved_seed.candidate_market_name,
        "market_path": resolved_seed.market_path,
        "upstream_batch_id": resolved_seed.upstream_batch_id,
        "upstream_status": resolved_seed.upstream_status,
    }


def run_once(args: argparse.Namespace, *, replay_attempted: bool = False, replay_result: dict[str, Any] | None = None) -> tuple[int, dict[str, Any], Path]:
    context = resolve_context_from_namespace(args, require_direction_id=False)
    output_dir = output_dir_from_namespace(args)
    log_dir = log_dir_from_namespace(args)
    download_dir = safe_download_dir(args)
    raw_artifact_path = ensure_within_repo(output_dir / BENCHMARK_RAW_ARTIFACT, "raw_artifact_path")
    poll_interval_seconds = ensure_min_poll_interval(args.poll_interval_seconds)
    summary: dict[str, Any] = {
        "timestamp": iso_now(),
        "module": "benchmark_export",
        "status": "BLOCKED",
        "reason_code": "",
        "message": "",
        "context_source": context.context_source,
        "运行名称": context.run_name,
        "方向ID": context.direction_id,
        "方向词": context.keyword,
        "站点": context.site,
        "attempted_url": BENCHMARK_URL,
        "final_url": "",
        "page_title": "",
        "query_keyword": "",
        "candidate_market_name": "",
        "market_path": "",
        "applied_filters": {},
        "seed_source_step": "",
        "seed_source_gate_path": "",
        "raw_artifact_path": "",
        "raw_item_count": 0,
        "dry_run": bool(args.dry_run),
        "headless": bool(args.headless),
        "download_dir": str(download_dir),
        "workbook_download_path": "",
        "matched_task_name": "",
        "matched_status_value": "",
        "execution_mode": "",
        "execution_warning": "",
        "auth_incident_path": "",
        "auth_surface_family": "",
        "auth_replay_available": False,
        "auth_replay_snippet_path": "",
        "auth_owner_recording_drop_path": "",
        "auth_replay_attempted": replay_attempted,
        "auth_replay_result": replay_result or {},
    }

    browser = None
    context_browser = None
    try:
        if args.max_wait_seconds <= 0:
            raise BenchmarkExportFlowError("--max-wait-seconds must be > 0.", "INVALID_MAX_WAIT")

        seed = resolve_seed(args, context)
        summary["query_keyword"] = seed["seed_keyword"]
        summary["candidate_market_name"] = seed["candidate_market_name"]
        summary["market_path"] = seed["market_path"]
        summary["seed_source_step"] = seed["source_step"]
        summary["seed_source_gate_path"] = seed["source_gate_path"]
        auth_context = {
            "context": {
                "run_name": context.run_name,
                "direction_id": context.direction_id,
                "keyword": context.keyword,
                "site": context.site,
                "days": context.days,
                "context_source": context.context_source,
            },
            "seed": seed,
            "download_dir": str(download_dir),
            "requested_execution_mode": args.execution_mode,
        }

        with sync_playwright() as playwright:
            context_browser, browser, execution_mode, warning = launch_context(playwright, args)
            summary["execution_mode"] = execution_mode
            summary["execution_warning"] = warning
            try:
                page = context_browser.pages[0] if context_browser.pages else context_browser.new_page()
                export_log_page = context_browser.new_page()
                export_log_page.goto(EXPORT_LOG_URL, wait_until="domcontentloaded", timeout=90000)
                wait_for_page_open(export_log_page)
                guard_page(export_log_page, "stage_b_export_log_baseline", preserve_texts=("下载", "我的导出"))
                if login_required(export_log_page):
                    fail_closed_on_auth(
                        export_log_page,
                        summary,
                        "stage_b_export_log_baseline",
                        "SellerSprite export-log page requires authentication.",
                        EXPORT_LOG_URL,
                        auth_context,
                    )
                baseline_tasks = collect_export_tasks(export_log_page)
                baseline_task_names = {str(task.get("task_name", "")).strip() for task in baseline_tasks if str(task.get("task_name", "")).strip()}
                record_step(
                    summary,
                    "stage_b_export_log_baseline",
                    "PASS",
                    page=export_log_page,
                    extra={"baseline_task_count": len(baseline_task_names)},
                )

                page.goto(BENCHMARK_URL, wait_until="domcontentloaded", timeout=90000)
                wait_for_page_open(page)
                summary["page_title"] = page.title()
                guard = guard_page(page, "stage_a_open_query_surface", preserve_texts=("立即查询", "导出", "导出明细"))
                if login_required(page):
                    fail_closed_on_auth(
                        page,
                        summary,
                        "stage_a_open_query_surface",
                        "SellerSprite benchmark page requires authentication.",
                        BENCHMARK_URL,
                        auth_context,
                    )
                ensure_query_surface(page, summary)
                summary["applied_filters"] = configure_filters(page, context.site, context.days)
                trigger_query(page, seed["seed_keyword"], summary)
                summary["final_url"] = page.url

                if login_required(page):
                    fail_closed_on_auth(
                        page,
                        summary,
                        "stage_a_query_results",
                        "SellerSprite benchmark query redirected to login.",
                        BENCHMARK_URL,
                        auth_context,
                    )
                if result_rows(page).count() <= 0:
                    screenshot_path = capture_screenshot(page, "stage_a_query_results-missing", SCREENSHOT_DIR)
                    record_step(
                        summary,
                        "stage_a_query_results",
                        "FAIL",
                        page=page,
                        reason_code=BENCHMARK_RESULT_TABLE_NOT_VISIBLE,
                        screenshot_path=screenshot_path,
                        extra={"guard": guard},
                    )
                    raise BenchmarkExportFlowError("Benchmark result table did not appear after query.", BENCHMARK_RESULT_TABLE_NOT_VISIBLE)

                record_step(
                    summary,
                    "stage_a_query_results",
                    "PASS",
                    page=page,
                    extra={"row_count": result_rows(page).count(), "guard": guard},
                )

                select_result_row(page, args.result_row_index, summary)

                if args.dry_run:
                    summary["status"] = "PASS"
                    summary["reason_code"] = "DRY_RUN_ONLY"
                    summary["message"] = "Dry-run validated query surface, row selection, and export trigger visibility without dispatching live export."
                    export_button = locate_export_button(page)
                    summary["export_button_state"] = export_button.evaluate("(node) => ({text: (node.innerText || '').trim(), disabled: !!node.disabled, className: node.className})") if export_button else {}
                    persist_run_summary(log_dir, "latest_benchmark_export_run.json", "benchmark_export_runs.jsonl", summary)
                    return 0, summary, log_dir

                export_triggered_ts = time.time()
                _follow_action = trigger_export(page, summary)
                expected_prefix = expected_task_prefix(context.site, context.days)

                deadline = time.monotonic() + args.max_wait_seconds
                matched_task: dict[str, Any] | None = None
                while time.monotonic() <= deadline:
                    export_log_page.goto(EXPORT_LOG_URL, wait_until="domcontentloaded", timeout=90000)
                    wait_for_page_open(export_log_page)
                    guard_page(export_log_page, "stage_b_poll_export_log", preserve_texts=("下载", "我的导出"))
                    if login_required(export_log_page):
                        fail_closed_on_auth(
                            export_log_page,
                            summary,
                            "stage_b_poll_export_log",
                            "SellerSprite export-log page redirected to login during polling.",
                            EXPORT_LOG_URL,
                            auth_context,
                        )
                    tasks = collect_export_tasks(export_log_page)
                    matched_task = find_best_export_task(tasks, baseline_task_names, expected_prefix, export_triggered_ts)
                    if matched_task is None:
                        record_step(
                            summary,
                            "stage_b_poll_export_log",
                            "WAIT",
                            page=export_log_page,
                            extra={"status_value": "", "matched_task_name": "", "visible_task_count": len(tasks)},
                        )
                        time.sleep(poll_interval_seconds)
                        continue

                    summary["matched_task_name"] = matched_task["task_name"]
                    summary["matched_status_value"] = matched_task["status_value"]
                    status_kind = export_status_kind(str(matched_task.get("status_value", "")))
                    if status_kind == "DONE":
                        record_step(
                            summary,
                            "stage_b_poll_export_log",
                            "PASS",
                            page=export_log_page,
                            extra={
                                "matched_task_name": matched_task["task_name"],
                                "status_value": matched_task["status_value"],
                                "source_name": matched_task["source_name"],
                            },
                        )
                        workbook_path, file_info = download_task_workbook(
                            export_log_page,
                            matched_task,
                            download_dir,
                            args.download_timeout_seconds * 1000,
                            summary,
                        )
                        summary["workbook_download_path"] = str(workbook_path)
                        try:
                            sheet_name, headers, items = parse_workbook_rows(workbook_path)
                        except Exception as exc:
                            raise BenchmarkExportFlowError(
                                f"Downloaded benchmark workbook could not be parsed into raw artifact: {workbook_path.name}",
                                EXPORT_WORKBOOK_PARSE_FAILED,
                            ) from exc
                        raw_artifact = build_raw_artifact(context, seed, workbook_path, sheet_name, headers, items)
                        raw_artifact["query_url"] = page.url
                        raw_artifact["page_title"] = page.title()
                        raw_artifact["applied_filters"] = summary["applied_filters"]
                        raw_artifact["export_log_task"] = {
                            "task_name": matched_task["task_name"],
                            "source_name": matched_task["source_name"],
                            "generated_at": matched_task["generated_at"],
                            "download_at": matched_task["download_at"],
                            "status_value": matched_task["status_value"],
                            "download_href": matched_task["download_href"],
                        }
                        output_dir.mkdir(parents=True, exist_ok=True)
                        write_json_atomic(raw_artifact_path, raw_artifact)

                        summary["status"] = "PASS"
                        summary["reason_code"] = "PASS"
                        summary["message"] = "Benchmark export workbook was downloaded, validated, parsed, and persisted as STEP4 raw artifact."
                        summary["raw_artifact_path"] = str(raw_artifact_path)
                        summary["raw_item_count"] = len(items)
                        summary["response_meta"] = {
                            "sheet_name": sheet_name,
                            "row_count": clean_number(len(items)),
                            "header_count": clean_number(len(headers)),
                        }
                        summary["download_file"] = file_info
                        break
                    if status_kind == "FAIL":
                        raise BenchmarkExportFlowError(
                            f"Benchmark export task failed with status `{matched_task['status_value']}`.",
                            EXPORT_LOG_STATUS_FAILED,
                        )

                    record_step(
                        summary,
                        "stage_b_poll_export_log",
                        "WAIT",
                        page=export_log_page,
                        extra={
                            "matched_task_name": matched_task["task_name"],
                            "status_value": matched_task["status_value"],
                            "source_name": matched_task["source_name"],
                        },
                    )
                    time.sleep(poll_interval_seconds)

                if summary.get("status") != "PASS":
                    if matched_task is None:
                        raise BenchmarkExportFlowError(
                            "No new benchmark export task could be locked in 我的导出 before timeout.",
                            EXPORT_LOG_TASK_NOT_FOUND,
                        )
                    raise BenchmarkExportFlowError(
                        f"Benchmark export task `{matched_task['task_name']}` did not complete before timeout.",
                        EXPORT_LOG_STATUS_TIMEOUT,
                    )
            finally:
                context_browser.close()
                if browser is not None:
                    browser.close()
    except BenchmarkChainError as exc:
        summary["status"] = "BLOCKED"
        summary["reason_code"] = exc.reason_code
        summary["message"] = str(exc)
    except BenchmarkExportFlowError as exc:
        summary["status"] = "BLOCKED"
        summary["reason_code"] = exc.reason_code
        summary["message"] = str(exc)
    except Exception as exc:
        summary["status"] = "BLOCKED"
        summary["reason_code"] = "BENCHMARK_EXPORT_UNHANDLED_ERROR"
        summary["message"] = str(exc)

    if summary["status"] != "PASS":
        try:
            if context_browser is not None and context_browser.pages:
                screenshot_path = capture_screenshot(context_browser.pages[0], "benchmark-export-flow-failure", SCREENSHOT_DIR)
                summary.setdefault("screenshots", []).append(screenshot_path)
        except Exception:
            pass

    persist_run_summary(log_dir, "latest_benchmark_export_run.json", "benchmark_export_runs.jsonl", summary)
    return (0 if summary["status"] == "PASS" else 2), summary, log_dir


def main() -> int:
    args = parse_args()
    exit_code, summary, log_dir = run_once(args)
    if exit_code != 0 and summary_requests_auth_replay(summary):
        replay_result = perform_registered_login_replay(
            surface_family=str(summary.get("auth_surface_family", "")).strip(),
            module_name="benchmark_export",
            trigger_reason_code=str(summary.get("reason_code", "")).strip(),
            trigger_summary=summary,
        )
        if replay_result.get("status") == "PASS":
            args.execution_mode = str(replay_result.get("execution_mode_override", "")).strip() or "storage_state"
            exit_code, summary, log_dir = run_once(args, replay_attempted=True, replay_result=replay_result)
        else:
            summary["auth_replay_attempted"] = True
            summary["auth_replay_result"] = replay_result
            persist_run_summary(log_dir, "latest_benchmark_export_run.json", "benchmark_export_runs.jsonl", summary)
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return exit_code


if __name__ == "__main__":
    raise SystemExit(main())
