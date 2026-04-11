from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from keyword_chain_common import (
    KeywordChainError,
    ensure_within_repo,
    iso_now,
    output_dir_from_namespace,
    resolve_context_from_namespace,
    traffic_cost_index_from_bid,
    write_json_atomic,
)


ROOT = Path(__file__).resolve().parents[1]
RAW_FILE_NAME = "keyword_research_raw.json"

HEADER_MAP = {
    "站点": "site",
    "月份": "month",
    "关键词": "keyword",
    "ABA排名": "search_rank",
    "月搜索量": "monthly_searches",
    "搜索增长率": "growth_pct",
    "点击总占比": "click_concentration_pct",
    "PPC价格": "ppc_bid_usd",
    "所属类目": "main_category",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Parse a SellerSprite KeywordHistory workbook into the STEP2 keyword_research raw artifact format.",
    )
    parser.add_argument("--workbook", required=True)
    parser.add_argument("--context-row-index", type=int, default=1)
    parser.add_argument("--run-name", default=None)
    parser.add_argument("--direction-id", default=None)
    parser.add_argument("--keyword", default=None)
    parser.add_argument("--category-hint", default=None)
    parser.add_argument("--site", default=None)
    parser.add_argument("--days", type=int, default=None)
    parser.add_argument("--sample-top-n", type=int, default=None)
    parser.add_argument("--output-dir", default=None)
    return parser.parse_args()


def first_number(text: Any) -> str:
    match = re.search(r"-?\d+(?:\.\d+)?", str(text or "").replace(",", ""))
    return match.group(0) if match else ""


def normalize_pct_value(value: Any) -> str:
    raw = first_number(value)
    if not raw:
        return ""
    numeric = float(raw)
    if abs(numeric) <= 5:
        numeric *= 100.0
    return f"{numeric:.4f}".rstrip("0").rstrip(".")


def select_sheet(workbook_path: Path):
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    for name in workbook.sheetnames:
        if "Notes" in name:
            continue
        worksheet = workbook[name]
        rows = worksheet.iter_rows(min_row=1, max_row=3, values_only=True)
        header = next(rows, None)
        if not header:
            continue
        header_values = [str(value or "").strip() for value in header]
        if "关键词" in header_values and "月搜索量" in header_values and "搜索增长率" in header_values:
            return workbook, worksheet, header_values, name
    workbook.close()
    raise KeywordChainError(
        f"KeywordHistory workbook did not contain a parsable data sheet: {workbook_path}",
        "KEYWORD_RESEARCH_WORKBOOK_SHEET_MISSING",
    )


def row_dict(header: list[str], values: list[Any]) -> dict[str, Any]:
    return {header[index]: values[index] if index < len(values) else "" for index in range(len(header))}


def build_rows(workbook_path: Path, context, raw_artifact_path: Path) -> tuple[list[dict[str, Any]], str, dict[str, Any]]:
    workbook, worksheet, header_values, sheet_name = select_sheet(workbook_path)
    try:
        first_data_values = None
        for values in worksheet.iter_rows(min_row=2, values_only=True):
            if any(str(value or "").strip() for value in values):
                first_data_values = list(values)
                break
        if first_data_values is None:
            raise KeywordChainError(
                f"KeywordHistory workbook had no data rows: {workbook_path}",
                "KEYWORD_RESEARCH_WORKBOOK_NO_ROWS",
            )

        item = row_dict(header_values, first_data_values)
        keyword = str(item.get("关键词") or context.keyword or "").strip()
        if not keyword:
            raise KeywordChainError(
                f"KeywordHistory workbook latest row did not contain a keyword: {workbook_path}",
                "KEYWORD_RESEARCH_WORKBOOK_NO_ROWS",
            )

        ppc_bid = first_number(item.get("PPC价格"))
        rows = [
            {
                "source_module": "KeywordResearchExport",
                "keyword": keyword,
                "site": str(item.get("站点") or context.site or "").strip().upper(),
                "main_category": str(item.get("所属类目") or context.category_hint or "").strip(),
                "monthly_searches": first_number(item.get("月搜索量")),
                "search_frequency_rank": first_number(item.get("ABA排名")),
                "growth_pct": normalize_pct_value(item.get("搜索增长率")),
                "click_concentration_pct": normalize_pct_value(item.get("点击总占比")),
                "ppc_bid_usd": ppc_bid,
                "traffic_cost_index": traffic_cost_index_from_bid(ppc_bid),
                "captured_at": iso_now(),
                "source_query": context.keyword,
                "source_file": str(workbook_path),
                "history_month": str(item.get("月份") or "").strip(),
                "raw_cells": [str(value or "") for value in first_data_values],
            }
        ]
        meta = {
            "sheet_name": sheet_name,
            "header": header_values,
            "history_month": rows[0]["history_month"],
        }
        return rows, sheet_name, meta
    finally:
        workbook.close()


def build_raw_artifact(workbook_path: Path, context, raw_artifact_path: Path) -> dict[str, Any]:
    workbook_path = ensure_within_repo(workbook_path, "keyword_history_workbook")
    rows, sheet_name, meta = build_rows(workbook_path, context, raw_artifact_path)
    return {
        "module": "keyword_research",
        "source_type": "SELLERSPRITE_KEYWORD_HISTORY_EXPORT_WORKBOOK",
        "status": "PASS",
        "timestamp": iso_now(),
        "url": "https://www.sellersprite.com/v3/keyword-miner",
        "title": f"KeywordHistory workbook | {sheet_name}",
        "workbook_path": str(workbook_path),
        "sheet_name": sheet_name,
        "context": {
            "run_name": context.run_name,
            "direction_id": context.direction_id,
            "keyword": context.keyword,
            "site": context.site,
            "category_hint": context.category_hint,
        },
        "workbook_meta": meta,
        "rows": rows,
    }


def main() -> int:
    args = parse_args()
    context = resolve_context_from_namespace(args, require_direction_id=False)
    output_dir = output_dir_from_namespace(args)
    raw_artifact_path = ensure_within_repo(output_dir / RAW_FILE_NAME, "raw_artifact_path")
    workbook_path = Path(args.workbook).expanduser()
    if not workbook_path.is_absolute():
        workbook_path = ROOT / workbook_path
    raw_artifact = build_raw_artifact(workbook_path, context, raw_artifact_path)
    output_dir.mkdir(parents=True, exist_ok=True)
    write_json_atomic(raw_artifact_path, raw_artifact)
    print(json.dumps({"status": "PASS", "raw_artifact_path": str(raw_artifact_path), "row_count": len(raw_artifact["rows"])}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
