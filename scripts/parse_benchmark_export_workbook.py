from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from benchmark_chain_common import (
    BENCHMARK_RAW_ARTIFACT,
    BenchmarkChainError,
    ensure_within_repo,
    iso_now,
    output_dir_from_namespace,
    resolve_context_from_namespace,
    resolve_seed_from_step3,
    write_json_atomic,
)


HEADER_ALIASES = {
    "asin": ("ASIN",),
    "title": ("商品标题", "商品名称"),
    "brand": ("品牌",),
    "price": ("价格($)", "价格"),
    "rating": ("评分",),
    "reviews": ("评分数", "评论数"),
    "bsrRank": ("大类BSR", "小类BSR"),
    "parent": ("父ASIN",),
    "variations": ("变体数",),
    "sellerType": ("配送方式", "BuyBox类型"),
    "categoryPath": ("类目路径",),
    "categoryName": ("大类目",),
    "subcategoryName": ("小类目",),
    "sku": ("SKU",),
    "imageUrl": ("商品主图",),
    "detailUrl": ("商品详情页链接",),
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Parse a downloaded SellerSprite benchmark workbook into the raw JSON artifact expected by STEP4 builder.",
    )
    parser.add_argument("--workbook", required=True)
    parser.add_argument("--context-row-index", type=int, default=1)
    parser.add_argument("--run-name", default=None)
    parser.add_argument("--direction-id", default=None)
    parser.add_argument("--keyword", default=None)
    parser.add_argument("--category-hint", default=None)
    parser.add_argument("--site", default=None)
    parser.add_argument("--days", type=int, default=None)
    parser.add_argument("--sample-top-n", type=int, default=None)
    parser.add_argument("--max-candidate-samples", type=int, default=None)
    parser.add_argument("--seed-keyword", default=None)
    parser.add_argument("--seed-market-name", default=None)
    parser.add_argument("--step3-gate-csv", default=None)
    parser.add_argument("--step3-cleaned-csv", default=None)
    parser.add_argument("--output-dir", default=None)
    parser.add_argument("--output-json", default=None)
    return parser.parse_args()


def normalize_header(value: Any) -> str:
    return str(value or "").strip()


def numeric_or_none(value: Any) -> float | int | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    if text.endswith("%"):
        text = text[:-1].strip()
    try:
        number = float(text)
    except ValueError:
        return None
    return int(number) if number.is_integer() else number


def text_or_blank(value: Any) -> str:
    return str(value or "").strip()


def choose_sheet(workbook_path: Path):
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    for sheet_name in workbook.sheetnames:
        if sheet_name.casefold() == "notes":
            continue
        worksheet = workbook[sheet_name]
        headers = [normalize_header(cell) for cell in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        if "ASIN" in headers and ("商品标题" in headers or "商品名称" in headers):
            return workbook, worksheet, sheet_name, headers
    raise BenchmarkChainError(
        f"No benchmark workbook sheet with required headers was found: {workbook_path}",
        "BENCHMARK_WORKBOOK_HEADERS_MISSING",
    )


def header_index_map(headers: list[str]) -> dict[str, int]:
    return {header: index for index, header in enumerate(headers) if header}


def value_by_alias(row: tuple[Any, ...], indices: dict[str, int], aliases: tuple[str, ...]) -> Any:
    for alias in aliases:
        index = indices.get(alias)
        if index is None:
            continue
        if index >= len(row):
            continue
        return row[index]
    return None


def parse_workbook_rows(workbook_path: Path) -> tuple[str, list[str], list[dict[str, Any]]]:
    workbook, worksheet, sheet_name, headers = choose_sheet(workbook_path)
    indices = header_index_map(headers)
    items: list[dict[str, Any]] = []
    try:
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            asin = text_or_blank(value_by_alias(row, indices, HEADER_ALIASES["asin"])).upper()
            title = text_or_blank(value_by_alias(row, indices, HEADER_ALIASES["title"]))
            if not asin or not title:
                continue
            item = {
                "asin": asin,
                "title": title,
                "brand": text_or_blank(value_by_alias(row, indices, HEADER_ALIASES["brand"])),
                "price": numeric_or_none(value_by_alias(row, indices, HEADER_ALIASES["price"])),
                "rating": numeric_or_none(value_by_alias(row, indices, HEADER_ALIASES["rating"])),
                "reviews": numeric_or_none(value_by_alias(row, indices, HEADER_ALIASES["reviews"])),
                "bsrRank": numeric_or_none(value_by_alias(row, indices, HEADER_ALIASES["bsrRank"])),
                "parent": text_or_blank(value_by_alias(row, indices, HEADER_ALIASES["parent"])).upper(),
                "variations": numeric_or_none(value_by_alias(row, indices, HEADER_ALIASES["variations"])),
                "sellerType": text_or_blank(value_by_alias(row, indices, HEADER_ALIASES["sellerType"])),
                "categoryPath": text_or_blank(value_by_alias(row, indices, HEADER_ALIASES["categoryPath"])),
                "categoryName": text_or_blank(value_by_alias(row, indices, HEADER_ALIASES["categoryName"])),
                "subcategoryName": text_or_blank(value_by_alias(row, indices, HEADER_ALIASES["subcategoryName"])),
                "sku": text_or_blank(value_by_alias(row, indices, HEADER_ALIASES["sku"])),
                "imageUrl": text_or_blank(value_by_alias(row, indices, HEADER_ALIASES["imageUrl"])),
                "asinUrl": text_or_blank(value_by_alias(row, indices, HEADER_ALIASES["detailUrl"])),
            }
            items.append(item)
    finally:
        workbook.close()

    if not items:
        raise BenchmarkChainError(
            f"Workbook opened but no usable competitor rows were parsed: {workbook_path}",
            "BENCHMARK_WORKBOOK_EMPTY",
        )
    return sheet_name, headers, items


def build_seed_context(args: argparse.Namespace, context) -> dict[str, Any]:
    if args.seed_keyword:
        seed_keyword = text_or_blank(args.seed_keyword)
        if not seed_keyword:
            raise BenchmarkChainError("--seed-keyword cannot be blank.", "SEED_KEYWORD_BLANK")
        return {
            "source_step": "MANUAL_OVERRIDE",
            "source_gate_path": "",
            "source_cleaned_path": "",
            "seed_keyword": seed_keyword,
            "candidate_market_name": text_or_blank(args.seed_market_name) or seed_keyword,
            "market_path": "",
            "upstream_batch_id": "",
            "upstream_status": "MANUAL_OVERRIDE",
        }

    resolved_seed = resolve_seed_from_step3(context, args.step3_gate_csv, args.step3_cleaned_csv)
    return {
        "source_step": resolved_seed.source_step,
        "source_gate_path": resolved_seed.source_gate_path,
        "source_cleaned_path": resolved_seed.source_cleaned_path,
        "seed_keyword": resolved_seed.seed_keyword,
        "candidate_market_name": resolved_seed.candidate_market_name,
        "market_path": resolved_seed.market_path,
        "upstream_batch_id": resolved_seed.upstream_batch_id,
        "upstream_status": resolved_seed.upstream_status,
    }


def build_raw_artifact(context, seed_context: dict[str, Any], workbook_path: Path, sheet_name: str, headers: list[str], items: list[dict[str, Any]]) -> dict[str, Any]:
    return {
        "module": "benchmark_export",
        "status": "PASS",
        "timestamp": iso_now(),
        "source_type": "SELLERSPRITE_EXPORT_WORKBOOK",
        "page_title": "查竞品|卖家精灵",
        "query_url": "",
        "workbook_path": str(workbook_path),
        "workbook_sheet_name": sheet_name,
        "workbook_headers": headers,
        "context": {
            "run_name": context.run_name,
            "direction_id": context.direction_id,
            "keyword": context.keyword,
            "category_hint": context.category_hint,
            "site": context.site,
            "days": context.days,
            "sample_top_n": context.sample_top_n,
            "max_candidate_samples": context.max_candidate_samples,
        },
        "seed_context": seed_context,
        "response_meta": {
            "sheet_name": sheet_name,
            "row_count": len(items),
            "header_count": len(headers),
        },
        "items": items,
    }


def main() -> int:
    args = parse_args()
    context = resolve_context_from_namespace(args, require_direction_id=False)
    workbook_path = Path(args.workbook).expanduser()
    if not workbook_path.is_absolute():
        workbook_path = Path.cwd() / workbook_path
    workbook_path = ensure_within_repo(workbook_path, "benchmark_workbook")
    if not workbook_path.exists():
        raise SystemExit(f"Workbook is missing: {workbook_path}")

    output_dir = output_dir_from_namespace(args)
    output_dir.mkdir(parents=True, exist_ok=True)
    output_json = Path(args.output_json).expanduser() if args.output_json else output_dir / BENCHMARK_RAW_ARTIFACT
    if not output_json.is_absolute():
        output_json = Path.cwd() / output_json
    output_json = ensure_within_repo(output_json, "output_json")

    seed_context = build_seed_context(args, context)
    sheet_name, headers, items = parse_workbook_rows(workbook_path)
    artifact = build_raw_artifact(context, seed_context, workbook_path, sheet_name, headers, items)
    write_json_atomic(output_json, artifact)
    print(
        json.dumps(
            {
                "status": "PASS",
                "reason_code": "PASS",
                "workbook_path": str(workbook_path),
                "output_json": str(output_json),
                "sheet_name": sheet_name,
                "row_count": len(items),
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
