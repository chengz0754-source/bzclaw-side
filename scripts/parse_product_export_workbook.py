from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from benchmark_chain_common import BenchmarkChainError, ensure_within_repo, iso_now


TITLE_HEADERS = ("商品标题", "商品名称", "标题", "Product Name", "Title")
ASIN_HEADERS = ("ASIN",)
HEADER_ALIASES = {
    "rank": ("序号", "排名", "#"),
    "title": TITLE_HEADERS,
    "asin": ASIN_HEADERS,
    "brand": ("品牌",),
    "parent_asin": ("父ASIN", "Parent ASIN"),
    "price_text": ("价格($)", "价格", "售价"),
    "rating_text": ("评分",),
    "review_text": ("评分数", "评论数", "Review Count"),
    "bsr_text": ("大类BSR", "小类BSR", "BSR"),
    "sales_text": ("销量(父)", "月销量", "销量"),
    "sales_amount_text": ("销售额", "月销售额", "销量额"),
    "child_sales_text": ("子体销量", "销量(子)"),
    "variation_text": ("变体数",),
    "gross_profit_text": ("毛利率", "毛利"),
    "launch_text": ("上架时间", "首次上架时间"),
    "delivery_text": ("配送", "配送方式", "卖家类型"),
    "category_path": ("类目路径", "中文类目名", "品类路径"),
    "category_name": ("大类目", "类目", "一级类目"),
    "subcategory_name": ("小类目", "子类目", "二级类目"),
    "detail_url": ("商品详情页链接", "商品详情链接", "商品链接", "ASIN链接"),
    "market_analysis_url": ("市场分析链接", "市场分析URL"),
    "image_url": ("商品主图", "主图"),
}


def normalize_header(value: Any) -> str:
    return str(value or "").strip()


def text_or_blank(value: Any) -> str:
    return str(value or "").strip()


def choose_sheet(workbook_path: Path):
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    for sheet_name in workbook.sheetnames:
        if sheet_name.casefold() == "notes":
            continue
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(min_row=1, max_row=1, values_only=True)
        headers = [normalize_header(cell) for cell in next(rows)]
        if "ASIN" in headers and any(alias in headers for alias in TITLE_HEADERS):
            return workbook, worksheet, sheet_name, headers
    workbook.close()
    raise BenchmarkChainError(
        f"No product export workbook sheet with required title/ASIN headers was found: {workbook_path}",
        "PRODUCT_WORKBOOK_HEADERS_MISSING",
    )


def header_index_map(headers: list[str]) -> dict[str, int]:
    return {header: idx for idx, header in enumerate(headers) if header}


def value_by_alias(row: tuple[Any, ...], indices: dict[str, int], aliases: tuple[str, ...]) -> Any:
    for alias in aliases:
        index = indices.get(alias)
        if index is None:
            continue
        if index >= len(row):
            continue
        return row[index]
    return None


def trailing_market_name(category_path: str, category_name: str, subcategory_name: str) -> str:
    if subcategory_name:
        return subcategory_name
    if category_name:
        return category_name
    parts = [segment.strip() for segment in str(category_path or "").split(">") if segment.strip()]
    return parts[-1] if parts else ""


def parse_workbook_rows(workbook_path: Path) -> tuple[str, list[str], list[dict[str, Any]]]:
    workbook, worksheet, sheet_name, headers = choose_sheet(workbook_path)
    indices = header_index_map(headers)
    items: list[dict[str, Any]] = []
    try:
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            asin = text_or_blank(value_by_alias(row, indices, HEADER_ALIASES["asin"])).upper()
            title = text_or_blank(value_by_alias(row, indices, HEADER_ALIASES["title"]))
            if not asin or not title:
                continue
            category_path = text_or_blank(value_by_alias(row, indices, HEADER_ALIASES["category_path"]))
            category_name = text_or_blank(value_by_alias(row, indices, HEADER_ALIASES["category_name"]))
            subcategory_name = text_or_blank(value_by_alias(row, indices, HEADER_ALIASES["subcategory_name"]))
            items.append(
                {
                    "rank": text_or_blank(value_by_alias(row, indices, HEADER_ALIASES["rank"])),
                    "title": title,
                    "asin": asin,
                    "brand": text_or_blank(value_by_alias(row, indices, HEADER_ALIASES["brand"])),
                    "parent_asin": text_or_blank(value_by_alias(row, indices, HEADER_ALIASES["parent_asin"])).upper(),
                    "price_text": text_or_blank(value_by_alias(row, indices, HEADER_ALIASES["price_text"])),
                    "rating_text": text_or_blank(value_by_alias(row, indices, HEADER_ALIASES["rating_text"])),
                    "review_text": text_or_blank(value_by_alias(row, indices, HEADER_ALIASES["review_text"])),
                    "bsr_text": text_or_blank(value_by_alias(row, indices, HEADER_ALIASES["bsr_text"])),
                    "sales_text": text_or_blank(value_by_alias(row, indices, HEADER_ALIASES["sales_text"])),
                    "sales_amount_text": text_or_blank(value_by_alias(row, indices, HEADER_ALIASES["sales_amount_text"])),
                    "child_sales_text": text_or_blank(value_by_alias(row, indices, HEADER_ALIASES["child_sales_text"])),
                    "variation_text": text_or_blank(value_by_alias(row, indices, HEADER_ALIASES["variation_text"])),
                    "gross_profit_text": text_or_blank(value_by_alias(row, indices, HEADER_ALIASES["gross_profit_text"])),
                    "launch_text": text_or_blank(value_by_alias(row, indices, HEADER_ALIASES["launch_text"])),
                    "delivery_text": text_or_blank(value_by_alias(row, indices, HEADER_ALIASES["delivery_text"])),
                    "category_path": category_path,
                    "candidate_market_name": trailing_market_name(category_path, category_name, subcategory_name),
                    "product_source_url": text_or_blank(value_by_alias(row, indices, HEADER_ALIASES["detail_url"])),
                    "market_analysis_url": text_or_blank(value_by_alias(row, indices, HEADER_ALIASES["market_analysis_url"])),
                    "image_url": text_or_blank(value_by_alias(row, indices, HEADER_ALIASES["image_url"])),
                    "source_file": str(workbook_path),
                }
            )
    finally:
        workbook.close()

    if not items:
        raise BenchmarkChainError(
            f"Workbook opened but no usable product rows were parsed: {workbook_path}",
            "PRODUCT_WORKBOOK_EMPTY",
        )
    return sheet_name, headers, items


def build_raw_artifact(context, workbook_path: Path, sheet_name: str, headers: list[str], items: list[dict[str, Any]], query_url: str, page_title: str) -> dict[str, Any]:
    return {
        "module": "product_research",
        "status": "PASS",
        "timestamp": iso_now(),
        "source_type": "SELLERSPRITE_PRODUCT_EXPORT_WORKBOOK",
        "page_title": page_title,
        "query_url": query_url,
        "workbook_path": str(workbook_path),
        "workbook_sheet_name": sheet_name,
        "workbook_headers": headers,
        "context": {
            "run_name": context.run_name,
            "direction_id": context.direction_id,
            "keyword": context.keyword,
            "category_hint": context.category_hint,
            "site": context.site,
            "days": context.days,
            "sample_top_n": context.sample_top_n,
            "max_candidate_samples": context.max_candidate_samples,
        },
        "response_meta": {
            "sheet_name": sheet_name,
            "row_count": len(items),
            "header_count": len(headers),
        },
        "items": items,
    }


def main() -> int:
    import argparse

    parser = argparse.ArgumentParser(description="Parse a SellerSprite product export workbook into normalized JSON rows.")
    parser.add_argument("--workbook", required=True)
    args = parser.parse_args()

    workbook_path = Path(args.workbook).expanduser()
    if not workbook_path.is_absolute():
        workbook_path = Path.cwd() / workbook_path
    workbook_path = ensure_within_repo(workbook_path, "product_workbook")
    sheet_name, headers, items = parse_workbook_rows(workbook_path)
    print(
        json.dumps(
            {
                "status": "PASS",
                "sheet_name": sheet_name,
                "header_count": len(headers),
                "row_count": len(items),
                "headers": headers,
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
