from __future__ import annotations

import argparse
import json
import re
import time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

from benchmark_chain_common import ensure_within_repo
from export_benchmark_competitors import (
    EXPORT_DOWNLOAD_BUTTON_NOT_VISIBLE,
    EXPORT_FILE_NOT_DOWNLOADED,
    EXPORT_LOG_STATUS_TIMEOUT,
    BenchmarkExportFlowError,
    collect_export_tasks,
    export_status_kind,
    launch_context,
    login_required,
    wait_before_click,
    wait_for_page_open,
)
from keyword_chain_common import ROOT


EXPORT_LOG_URL = "https://www.sellersprite.com/v2/export-log"
DOWNLOADS_ROOT = ROOT / "runs" / "manual" / "15_product_exports"
PRODUCT_EXPORT_SOURCE_MARKERS = ("选产品", "Product")


PRODUCT_EXPORT_SOURCE_MARKERS = ("选产品", "Product")

PRODUCT_EXPORT_SOURCE_MARKERS = ("选产品", "Product")

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Debug-only product export-log poller locked by task name.")
    parser.add_argument("--site", default="US")
    parser.add_argument("--days", type=int, default=30)
    parser.add_argument("--max-wait-seconds", type=int, default=300)
    parser.add_argument("--poll-interval-seconds", type=int, default=10)
    parser.add_argument("--download-timeout-seconds", type=int, default=90)
    parser.add_argument("--download-dir", default=None)
    parser.add_argument("--headless", action="store_true")
    parser.add_argument("--execution-mode", choices=("auto", "persistent", "storage_state"), default="auto")
    return parser.parse_args()


def ensure_min_poll_interval(seconds: int) -> int:
    return max(8, int(seconds))


def safe_download_dir(raw_value: str | None) -> Path:
    if raw_value:
        path = Path(raw_value).expanduser()
        if not path.is_absolute():
            path = ROOT / path
        return ensure_within_repo(path, "product_download_dir")
    return ensure_within_repo(DOWNLOADS_ROOT / time.strftime("%Y%m%d_%H%M%S"), "product_download_dir")


def expected_product_task_prefix(site: str, days: int) -> str:
    return f"Product-{str(site or '').strip().upper()}-Last-{int(days)}-days"


def safe_download_filename(task_name: str, suggested_filename: str) -> str:
    suffix = Path(suggested_filename or "").suffix.lower() or ".xlsx"
    stem = Path(suggested_filename or "").stem or task_name
    safe_stem = re.sub(r"[^A-Za-z0-9._-]+", "-", stem).strip("-") or "product-export"
    return f"{safe_stem}{suffix}"


def validate_downloaded_workbook(path: Path, expected_task_name: str) -> dict[str, Any]:
    if not path.exists():
        raise BenchmarkExportFlowError(f"Product export workbook was not saved: {path}", EXPORT_FILE_NOT_DOWNLOADED)
    if path.stat().st_size <= 0:
        raise BenchmarkExportFlowError(f"Product export workbook is empty: {path}", EXPORT_FILE_NOT_DOWNLOADED)
    if path.suffix.lower() not in {".xlsx", ".xls"}:
        raise BenchmarkExportFlowError(f"Unexpected product export suffix: {path.name}", EXPORT_FILE_NOT_DOWNLOADED)
    if str(expected_task_name or "").strip() and str(expected_task_name).casefold() not in path.stem.casefold():
        raise BenchmarkExportFlowError(
            f"Product export filename does not contain the locked task name `{expected_task_name}`: {path.name}",
            EXPORT_FILE_NOT_DOWNLOADED,
        )
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet_names = workbook.sheetnames
        workbook.close()
    except Exception as exc:
        raise BenchmarkExportFlowError(f"Product export workbook is unreadable: {path.name}", EXPORT_FILE_NOT_DOWNLOADED) from exc
    return {
        "path": str(path),
        "name": path.name,
        "size_bytes": path.stat().st_size,
        "sheet_names": sheet_names,
    }


def find_best_product_export_task(
    tasks: list[dict[str, Any]],
    baseline_task_names: set[str],
    expected_prefix: str,
    export_triggered_ts: float,
) -> dict[str, Any] | None:
    ranked: list[tuple[int, float, dict[str, Any]]] = []
    for task in tasks:
        task_name = str(task.get("task_name", "")).strip()
        source_name = str(task.get("source_name", "")).strip()
        if not task_name:
            continue
        score = 0
        if expected_prefix and task_name.startswith(expected_prefix):
            score += 500
        if any(marker.casefold() in source_name.casefold() for marker in PRODUCT_EXPORT_SOURCE_MARKERS):
            score += 300
        if task_name not in baseline_task_names:
            score += 1000
        generated_ts = task.get("generated_ts")
        if isinstance(generated_ts, (int, float)):
            delta = abs(float(generated_ts) - export_triggered_ts)
            if delta <= 300:
                score += 200
            elif delta <= 900:
                score += 100
        if score <= 0:
            continue
        ranked.append((score, float(generated_ts or 0.0), task))

    if not ranked:
        return None
    ranked.sort(key=lambda item: (item[0], item[1]), reverse=True)
    return ranked[0][2]


def poll_product_export_task(
    page,
    baseline_task_names: set[str],
    site: str,
    days: int,
    export_triggered_ts: float,
    max_wait_seconds: int,
    poll_interval_seconds: int,
) -> dict[str, Any]:
    expected_prefix = expected_product_task_prefix(site, days)
    deadline = time.monotonic() + max_wait_seconds
    locked_task_name = ""

    while time.monotonic() <= deadline:
        page.goto(EXPORT_LOG_URL, wait_until="domcontentloaded", timeout=90000)
        wait_for_page_open(page)
        if login_required(page):
            raise BenchmarkExportFlowError("SellerSprite export-log redirected product polling to login.", "SELLERSPRITE_AUTH_REQUIRED")
        tasks = collect_export_tasks(page)
        if locked_task_name:
            matched = next((task for task in tasks if str(task.get("task_name", "")).strip() == locked_task_name), None)
        else:
            matched = find_best_product_export_task(tasks, baseline_task_names, expected_prefix, export_triggered_ts)
            if matched is not None:
                locked_task_name = str(matched.get("task_name", "")).strip()
        if matched is None:
            time.sleep(poll_interval_seconds)
            continue
        status_kind = export_status_kind(str(matched.get("status_value", "")))
        if status_kind == "DONE":
            return matched
        if status_kind == "FAIL":
            raise BenchmarkExportFlowError(
                f"Product export task `{locked_task_name or matched.get('task_name', '')}` failed with status `{matched.get('status_value', '')}`.",
                "PRODUCT_EXPORT_STATUS_FAILED",
            )
        time.sleep(poll_interval_seconds)
    raise BenchmarkExportFlowError("Product export task did not complete before timeout.", EXPORT_LOG_STATUS_TIMEOUT)


def download_product_export_task(page, task: dict[str, Any], download_dir: Path, timeout_ms: int) -> tuple[Path, dict[str, Any]]:
    task_name = str(task.get("task_name", "")).strip()
    matching_row = page.locator("tbody tr").filter(has_text=task_name).first
    try:
        matching_row.wait_for(state="visible", timeout=10000)
    except Exception as exc:
        raise BenchmarkExportFlowError("Locked product export task row is no longer visible.", EXPORT_DOWNLOAD_BUTTON_NOT_VISIBLE) from exc

    button = matching_row.locator("a.download-excel, button.download-excel, .download-excel").first
    try:
        button.wait_for(state="visible", timeout=10000)
    except Exception as exc:
        raise BenchmarkExportFlowError("Product export download button is not visible in the locked row.", EXPORT_DOWNLOAD_BUTTON_NOT_VISIBLE) from exc

    wait_before_click(page)
    with page.expect_download(timeout=timeout_ms) as download_info:
        button.click(timeout=10000)
    download = download_info.value
    download_dir.mkdir(parents=True, exist_ok=True)
    target_path = ensure_within_repo(download_dir / safe_download_filename(task_name, download.suggested_filename), "product_download_target")
    download.save_as(str(target_path))
    file_info = validate_downloaded_workbook(target_path, task_name)
    return target_path, file_info


def main() -> int:
    args = parse_args()
    download_dir = safe_download_dir(args.download_dir)
    summary: dict[str, Any] = {
        "module": "temp_product_export_log_record",
        "status": "BLOCKED",
        "reason_code": "",
        "message": "",
        "download_dir": str(download_dir),
    }

    browser = None
    context_browser = None
    try:
        with sync_playwright() as playwright:
            context_browser, browser, execution_mode, warning = launch_context(
                playwright,
                args,
                runtime_replay_surface_family="SELLERSPRITE_EXPORT_LOG_AUTH",
            )
            summary["execution_mode"] = execution_mode
            if warning:
                summary["execution_warning"] = warning
            page = context_browser.pages[0] if context_browser.pages else context_browser.new_page()
            page.goto(EXPORT_LOG_URL, wait_until="domcontentloaded", timeout=90000)
            wait_for_page_open(page)
            if login_required(page):
                raise BenchmarkExportFlowError("SellerSprite export-log requires authentication.", "SELLERSPRITE_AUTH_REQUIRED")
            tasks = collect_export_tasks(page)
            summary["visible_task_count"] = len(tasks)
            summary["task_preview"] = tasks[:10]
            summary["status"] = "PASS"
            summary["reason_code"] = "PASS"
            summary["message"] = "Current export-log tasks were collected without clicking the first button."
    except Exception as exc:
        summary["status"] = "BLOCKED"
        summary["reason_code"] = getattr(exc, "reason_code", "TEMP_PRODUCT_EXPORT_LOG_FAILED")
        summary["message"] = str(exc)
    finally:
        if context_browser is not None:
            try:
                context_browser.close()
            except Exception:
                pass
        if browser is not None:
            try:
                browser.close()
            except Exception:
                pass

    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0 if summary["status"] == "PASS" else 2


if __name__ == "__main__":
    raise SystemExit(main())
